from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, Request, HTTPException, Response, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import logging
import bcrypt
import jwt
import uuid
import stripe
import base64
import hashlib
import hmac as hmac_lib
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import io
import unicodedata
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, timedelta
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import padding as sym_padding
from cryptography.hazmat.backends import default_backend
from pydantic import BaseModel, Field
from typing import List, Optional

stripe.api_key = os.environ.get("STRIPE_SECRET_KEY", "")

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_ALGORITHM = "HS256"

def get_jwt_secret():
    return os.environ["JWT_SECRET"]

def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode("utf-8"), salt).decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(hours=24), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request):
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="No autenticado")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Token invalido")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalido")

# Pydantic models
class LoginRequest(BaseModel):
    email: str
    password: str

class NewsCreate(BaseModel):
    title: str
    content: str
    image_url: Optional[str] = ""
    source: Optional[str] = "web"
    category: Optional[str] = "general"

class NewsUpdate(BaseModel):
    title: Optional[str] = None
    content: Optional[str] = None
    image_url: Optional[str] = None
    source: Optional[str] = None
    category: Optional[str] = None

# Webhook model - receives posts from n8n/Make
class SocialPostWebhook(BaseModel):
    source: str  # "instagram" or "facebook"
    content: Optional[str] = ""
    image_url: Optional[str] = ""
    post_url: Optional[str] = ""
    author: Optional[str] = ""
    timestamp: Optional[str] = ""
    api_key: Optional[str] = ""

class TeamCreate(BaseModel):
    name: str
    category: str
    coach: Optional[str] = ""
    coach_ids: Optional[List[str]] = []    # user IDs with role=entrenador
    player_ids: Optional[List[str]] = []   # player IDs assigned to this team
    image_url: Optional[str] = ""
    description: Optional[str] = ""
    facility_id: Optional[str] = ""

class PlayerCreate(BaseModel):
    name: str
    surname: Optional[str] = ""
    position: Optional[str] = ""
    number: Optional[int] = None
    team_id: str
    image_url: Optional[str] = ""
    # Datos personales
    birthdate: Optional[str] = ""
    dni: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    postal_code: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    # Club
    jersey_size: Optional[str] = ""
    season: Optional[str] = "2025/2026"
    # Salud
    medical_notes: Optional[str] = ""
    blood_type: Optional[str] = ""
    # Familia
    family_id: Optional[str] = ""
    # Estado
    status: Optional[str] = "active"
    notes: Optional[str] = ""

class GuardianCreate(BaseModel):
    name: str
    surname: Optional[str] = ""
    dni: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    relationship: Optional[str] = "padre"
    player_ids: Optional[List[str]] = []
    family_id: Optional[str] = ""
    bank_iban: Optional[str] = ""
    notes: Optional[str] = ""

class GuardianUpdate(BaseModel):
    name: Optional[str] = None
    surname: Optional[str] = None
    dni: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    relationship: Optional[str] = None
    player_ids: Optional[List[str]] = None
    family_id: Optional[str] = None
    bank_iban: Optional[str] = None
    notes: Optional[str] = None

class MemberCreate(BaseModel):
    name: str
    surname: Optional[str] = ""
    dni: Optional[str] = ""
    birthdate: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    postal_code: Optional[str] = ""
    member_type: Optional[str] = "socio_adulto"
    bank_iban: Optional[str] = ""
    photo_url: Optional[str] = ""
    status: Optional[str] = "active"
    season: Optional[str] = "2025/2026"
    notes: Optional[str] = ""

class MemberUpdate(BaseModel):
    name: Optional[str] = None
    surname: Optional[str] = None
    dni: Optional[str] = None
    birthdate: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    postal_code: Optional[str] = None
    member_type: Optional[str] = None
    bank_iban: Optional[str] = None
    photo_url: Optional[str] = None
    status: Optional[str] = None
    season: Optional[str] = None
    notes: Optional[str] = None

class TrainingScheduleCreate(BaseModel):
    team_id: str
    day_of_week: str
    start_time: str
    end_time: str
    location: Optional[str] = ""
    notes: Optional[str] = ""

class FeeCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    amount: float
    currency: Optional[str] = "eur"
    fee_type: Optional[str] = "inscripcion"  # inscripcion | cuota_mensual | cuota_temporada | socio
    team_id: Optional[str] = ""
    member_type: Optional[str] = ""
    active: Optional[bool] = True

class FeeUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    amount: Optional[float] = None
    fee_type: Optional[str] = None
    team_id: Optional[str] = None
    member_type: Optional[str] = None
    active: Optional[bool] = None

class PaymentCreate(BaseModel):
    person_id: str
    person_type: str  # player | member
    fee_id: Optional[str] = ""
    amount: float
    concept: str
    status: Optional[str] = "pending"  # pending | paid | failed | refunded
    due_date: Optional[str] = ""
    notes: Optional[str] = ""

class PaymentUpdate(BaseModel):
    status: Optional[str] = None
    paid_at: Optional[str] = None
    notes: Optional[str] = None
    stripe_payment_intent_id: Optional[str] = None

# Public registration models (no auth needed)
class GuardianInfo(BaseModel):
    name: str
    surname: Optional[str] = ""
    dni: Optional[str] = ""
    phone: str
    email: str
    relationship: Optional[str] = "padre"

class PlayerRegistrationRequest(BaseModel):
    # Player data
    name: str
    surname: str
    birthdate: str
    dni: Optional[str] = ""
    phone: str
    email: str
    address: Optional[str] = ""
    city: Optional[str] = ""
    postal_code: Optional[str] = ""
    team_id: str
    position: Optional[str] = ""
    jersey_size: Optional[str] = ""
    medical_notes: Optional[str] = ""
    # Guardian (required if minor)
    guardian: Optional[GuardianInfo] = None
    # Payment
    fee_id: str
    payment_method: Optional[str] = "stripe"  # stripe | bank_transfer | redsys
    success_url: Optional[str] = ""
    cancel_url: Optional[str] = ""
    # RGPD — consentimiento (requerido)
    consent_gdpr: Optional[bool] = False
    consent_communications: Optional[bool] = False
    consent_guardian_gdpr: Optional[bool] = False  # para menores: consentimiento del tutor

class MemberRegistrationRequest(BaseModel):
    name: str
    surname: str
    birthdate: Optional[str] = ""
    dni: Optional[str] = ""
    phone: str
    email: str
    address: Optional[str] = ""
    city: Optional[str] = ""
    postal_code: Optional[str] = ""
    member_type: Optional[str] = "socio_adulto"
    bank_iban: Optional[str] = ""
    fee_id: str
    payment_method: Optional[str] = "stripe"  # stripe | bank_transfer | redsys
    success_url: Optional[str] = ""
    cancel_url: Optional[str] = ""
    # RGPD — consentimiento (requerido)
    consent_gdpr: Optional[bool] = False
    consent_communications: Optional[bool] = False

class MatchCreate(BaseModel):
    home_team: str
    away_team: str
    date: str
    time: str
    location: str
    category: Optional[str] = ""
    result: Optional[str] = ""
    status: Optional[str] = "upcoming"

class GalleryCreate(BaseModel):
    title: str
    image_url: str
    description: Optional[str] = ""
    category: Optional[str] = "general"
    section: Optional[str] = "galeria"  # hero|equipos|instalaciones|galeria|patrocinadores|noticias
    visible: Optional[bool] = True

class ContactCreate(BaseModel):
    name: str
    email: str
    phone: Optional[str] = ""
    message: str
    subject: Optional[str] = "general"

class SettingsUpdate(BaseModel):
    club_name: Optional[str] = None
    description: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    instagram_url: Optional[str] = None
    facebook_url: Optional[str] = None
    instagram_username: Optional[str] = None
    facebook_page: Optional[str] = None
    # Social feed integration
    facebook_embed_enabled: Optional[bool] = None
    facebook_page_url: Optional[str] = None
    instagram_embed_code: Optional[str] = None
    social_feed_provider: Optional[str] = None
    custom_embed_code: Optional[str] = None
    # Stripe (configurado por el club)
    stripe_secret_key: Optional[str] = None
    stripe_public_key: Optional[str] = None
    stripe_webhook_secret: Optional[str] = None
    stripe_enabled: Optional[bool] = None
    # Redsys TPV virtual
    redsys_merchant_code: Optional[str] = None
    redsys_terminal: Optional[str] = None
    redsys_secret_key: Optional[str] = None
    redsys_environment: Optional[str] = None   # test | production
    redsys_enabled: Optional[bool] = None
    # Transferencia bancaria
    bank_transfer_enabled: Optional[bool] = None
    bank_iban: Optional[str] = None
    bank_bic: Optional[str] = None
    bank_name: Optional[str] = None
    bank_holder: Optional[str] = None
    bank_transfer_instructions: Optional[str] = None
    # SEPA Direct Debit
    sepa_enabled: Optional[bool] = None
    sepa_creditor_id: Optional[str] = None
    sepa_creditor_name: Optional[str] = None
    sepa_creditor_iban: Optional[str] = None
    sepa_creditor_bic: Optional[str] = None
    # Email SMTP (cuenta propia del club)
    smtp_enabled: Optional[bool] = None
    smtp_host: Optional[str] = None
    smtp_port: Optional[int] = None
    smtp_user: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_use_tls: Optional[bool] = None
    smtp_from_name: Optional[str] = None
    # RGPD
    privacy_policy: Optional[str] = None

class SepaMandateCreate(BaseModel):
    person_id: str
    person_type: str      # player | member
    person_name: str
    debtor_iban: str
    debtor_bic: Optional[str] = ""
    mandate_ref: Optional[str] = ""   # auto-generated if empty
    signature_date: str               # YYYY-MM-DD
    sequence_type: Optional[str] = "FRST"  # FRST | RCUR | FNAL | OOFF
    amount: Optional[float] = None    # cuota del titular (sobrescribe default_amount en XML)
    notes: Optional[str] = ""

class SepaXmlRequest(BaseModel):
    mandate_ids: List[str]
    collection_date: str   # YYYY-MM-DD — fecha de cobro
    concept: str
    remesa_number: Optional[int] = None   # número de remesa (ej. 499)
    default_amount: Optional[float] = None  # importe por defecto si el mandato no tiene amount propio

class RedsysNotification(BaseModel):
    Ds_SignatureVersion: Optional[str] = ""
    Ds_MerchantParameters: Optional[str] = ""
    Ds_Signature: Optional[str] = ""

class FacilityCreate(BaseModel):
    name: str
    address: Optional[str] = ""
    city: Optional[str] = ""
    surface: Optional[str] = ""       # cesped | cesped_artificial | tartan | parquet | asfalto
    capacity: Optional[int] = None
    dimensions: Optional[str] = ""    # "110x70m"
    indoor: Optional[bool] = False
    team_ids: Optional[List[str]] = []
    photo_url: Optional[str] = ""
    notes: Optional[str] = ""

class FacilityUpdate(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    surface: Optional[str] = None
    capacity: Optional[int] = None
    dimensions: Optional[str] = None
    indoor: Optional[bool] = None
    team_ids: Optional[List[str]] = None
    photo_url: Optional[str] = None
    notes: Optional[str] = None

class SaleCreate(BaseModel):
    person_id: str
    person_type: str               # player | member
    product_id: Optional[str] = ""
    fee_id: Optional[str] = ""
    concept: str
    amount: float
    currency: Optional[str] = "eur"
    payment_method: Optional[str] = "pending"  # pending | cash | bank_transfer | stripe | redsys | sepa
    status: Optional[str] = "pending"          # pending | paid | cancelled | refunded
    due_date: Optional[str] = ""
    notes: Optional[str] = ""

class SaleUpdate(BaseModel):
    status: Optional[str] = None
    payment_method: Optional[str] = None
    paid_at: Optional[str] = None
    notes: Optional[str] = None
    sepa_mandate_id: Optional[str] = None

class BulkSaleRequest(BaseModel):
    player_ids: Optional[List[str]] = []
    member_ids: Optional[List[str]] = []
    product_id: Optional[str] = ""
    fee_id: Optional[str] = ""
    concept: str
    amount: float
    payment_method: Optional[str] = "pending"
    due_date: Optional[str] = ""
    notes: Optional[str] = ""

class ScheduleTemplateEventItem(BaseModel):
    type: Optional[str] = "entrenamiento"
    title: Optional[str] = ""
    team_id: Optional[str] = ""
    facility_id: Optional[str] = ""
    day_of_week: Optional[int] = 0  # 0=Mon, 6=Sun
    time: Optional[str] = "09:00"
    duration_min: Optional[int] = 90

class ScheduleTemplateCreate(BaseModel):
    name: str
    events: Optional[List[ScheduleTemplateEventItem]] = []
    # legacy single-event fields (kept for backward compat)
    team_id: Optional[str] = ""
    facility_id: Optional[str] = ""
    event_type: Optional[str] = "training"
    day_of_week: Optional[str] = ""
    start_time: Optional[str] = ""
    end_time: Optional[str] = ""
    repeat_weeks: Optional[int] = 1
    notes: Optional[str] = ""

class ScheduleEventCreate(BaseModel):
    title: Optional[str] = ""
    type: Optional[str] = "entrenamiento"   # new field
    team_id: Optional[str] = ""
    facility_id: Optional[str] = ""
    event_type: Optional[str] = "training"  # legacy
    date: str
    time: Optional[str] = ""               # new field (HH:MM)
    duration_min: Optional[int] = 90       # new field
    start_time: Optional[str] = ""         # legacy
    end_time: Optional[str] = ""           # legacy
    category: Optional[str] = ""
    opponent: Optional[str] = ""
    home_away: Optional[str] = "home"
    player_ids: Optional[List[str]] = []
    notes: Optional[str] = ""

class ProductCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    category: Optional[str] = "equipacion"  # equipacion | material | cuota | otro
    price: float = 0.0
    currency: Optional[str] = "eur"
    sku: Optional[str] = ""
    stock: Optional[int] = None
    active: Optional[bool] = True
    image_url: Optional[str] = ""
    is_recurring: Optional[bool] = False
    recurring_period: Optional[str] = "monthly"  # weekly | monthly | quarterly | yearly
    recurring_amount: Optional[float] = None

class ProductUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    price: Optional[float] = None
    sku: Optional[str] = None
    stock: Optional[int] = None
    active: Optional[bool] = None
    image_url: Optional[str] = None
    is_recurring: Optional[bool] = None
    recurring_period: Optional[str] = None
    recurring_amount: Optional[float] = None

class AdminUserCreate(BaseModel):
    email: str
    name: str
    password: str
    role: Optional[str] = "secretaria"  # admin | director | secretaria | entrenador

class AdminUserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    password: Optional[str] = None
    active: Optional[bool] = None

class CommListCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    filter_type: Optional[str] = "manual"  # manual | team | category | all_players | all_members | all_guardians
    filter_value: Optional[str] = ""
    recipient_ids: Optional[List[str]] = []
    recipient_type: Optional[str] = "player"  # player | member | guardian | user | mixed

class CommListUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    filter_type: Optional[str] = None
    filter_value: Optional[str] = None
    recipient_ids: Optional[List[str]] = None
    recipient_type: Optional[str] = None

class CommSendRequest(BaseModel):
    subject: str
    body_html: str
    list_id: Optional[str] = ""
    recipient_emails: Optional[List[str]] = []

# Create app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# --- AUTH ENDPOINTS ---
@api_router.post("/auth/login")
async def login(request_data: LoginRequest, response: Response):
    email = request_data.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(request_data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    return {"id": user_id, "email": user["email"], "name": user.get("name", ""), "role": user.get("role", "admin"), "token": access_token}

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return user

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Sesion cerrada"}

# --- NEWS ENDPOINTS ---
@api_router.get("/news")
async def get_news(limit: int = 50):
    news = await db.news.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return news

@api_router.post("/news")
async def create_news(data: NewsCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.news.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/news/{news_id}")
async def update_news(news_id: str, data: NewsUpdate, request: Request):
    await get_current_user(request)
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.news.update_one({"id": news_id}, {"$set": update_data})
    updated = await db.news.find_one({"id": news_id}, {"_id": 0})
    return updated

@api_router.delete("/news/{news_id}")
async def delete_news(news_id: str, request: Request):
    await get_current_user(request)
    await db.news.delete_one({"id": news_id})
    return {"message": "Noticia eliminada"}


# --- SOCIAL POSTS (from webhooks/automation) ---
@api_router.get("/social-posts")
async def get_social_posts(source: Optional[str] = None, limit: int = 8):
    query = {"source": source} if source else {}
    posts = await db.social_posts.find(query, {"_id": 0}).sort("posted_at", -1).to_list(limit)
    return posts

@api_router.post("/webhook/social-post")
async def receive_social_post(data: SocialPostWebhook):
    """Webhook endpoint - receives posts from n8n/Make/Zapier automation.
    No auth required (uses optional api_key for security).
    """
    webhook_key = os.environ.get("WEBHOOK_API_KEY", "")
    if webhook_key and data.api_key != webhook_key:
        raise HTTPException(status_code=403, detail="Invalid API key")
    
    doc = {
        "id": str(uuid.uuid4()),
        "source": data.source.lower(),
        "content": data.content or "",
        "image_url": data.image_url or "",
        "post_url": data.post_url or "",
        "author": data.author or "",
        "posted_at": data.timestamp or datetime.now(timezone.utc).isoformat(),
        "received_at": datetime.now(timezone.utc).isoformat()
    }
    await db.social_posts.insert_one(doc)
    return {"status": "ok", "id": doc["id"]}

@api_router.delete("/social-posts/{post_id}")
async def delete_social_post(post_id: str, request: Request):
    await get_current_user(request)
    await db.social_posts.delete_one({"id": post_id})
    return {"message": "Post eliminado"}


# --- TEAMS ENDPOINTS ---
@api_router.get("/teams")
async def get_teams():
    teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    return teams

@api_router.post("/teams")
async def create_team(data: TeamCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.teams.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/teams/{team_id}")
async def update_team(team_id: str, data: TeamCreate, request: Request):
    await get_current_user(request)
    update_data = data.model_dump()
    await db.teams.update_one({"id": team_id}, {"$set": update_data})
    updated = await db.teams.find_one({"id": team_id}, {"_id": 0})
    return updated

@api_router.delete("/teams/{team_id}")
async def delete_team(team_id: str, request: Request):
    await get_current_user(request)
    await db.teams.delete_one({"id": team_id})
    await db.players.delete_many({"team_id": team_id})
    return {"message": "Equipo eliminado"}

# --- PLAYERS ENDPOINTS ---
@api_router.get("/players")
async def get_players(team_id: Optional[str] = None, status: Optional[str] = None):
    query = {}
    if team_id:
        query["team_id"] = team_id
    if status:
        query["status"] = status
    players = await db.players.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return players

@api_router.get("/players/{player_id}")
async def get_player(player_id: str, request: Request):
    await get_current_user(request)
    player = await db.players.find_one({"id": player_id}, {"_id": 0})
    if not player:
        raise HTTPException(status_code=404, detail="Jugador no encontrado")
    return player

@api_router.post("/players")
async def create_player(data: PlayerCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    if not doc.get("family_id"):
        doc["family_id"] = str(uuid.uuid4())
    await db.players.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/players/{player_id}")
async def update_player(player_id: str, data: PlayerCreate, request: Request):
    await get_current_user(request)
    update_data = data.model_dump()
    await db.players.update_one({"id": player_id}, {"$set": update_data})
    updated = await db.players.find_one({"id": player_id}, {"_id": 0})
    return updated

@api_router.delete("/players/{player_id}")
async def delete_player(player_id: str, request: Request):
    await get_current_user(request)
    await db.players.delete_one({"id": player_id})
    return {"message": "Jugador eliminado"}

# --- GUARDIANS ENDPOINTS ---
@api_router.get("/guardians")
async def get_guardians(request: Request):
    await get_current_user(request)
    guardians = await db.guardians.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    return guardians

@api_router.get("/guardians/by-player/{player_id}")
async def get_guardians_by_player(player_id: str, request: Request):
    await get_current_user(request)
    guardians = await db.guardians.find({"player_ids": player_id}, {"_id": 0}).to_list(10)
    return guardians

@api_router.post("/guardians")
async def create_guardian(data: GuardianCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.guardians.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/guardians/{guardian_id}")
async def update_guardian(guardian_id: str, data: GuardianUpdate, request: Request):
    await get_current_user(request)
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.guardians.update_one({"id": guardian_id}, {"$set": update_data})
    updated = await db.guardians.find_one({"id": guardian_id}, {"_id": 0})
    return updated

@api_router.delete("/guardians/{guardian_id}")
async def delete_guardian(guardian_id: str, request: Request):
    await get_current_user(request)
    await db.guardians.delete_one({"id": guardian_id})
    return {"message": "Tutor eliminado"}

# --- MEMBERS (SOCIOS) ENDPOINTS ---
@api_router.get("/members")
async def get_members(request: Request, status: Optional[str] = None):
    await get_current_user(request)
    query = {"status": status} if status else {}
    members = await db.members.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return members

@api_router.get("/members/count")
async def get_members_count():
    count = await db.members.count_documents({})
    return {"count": count}

@api_router.post("/members")
async def create_member(data: MemberCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    count = await db.members.count_documents({})
    doc["member_number"] = f"S{count + 1:04d}"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.members.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/members/{member_id}")
async def update_member(member_id: str, data: MemberUpdate, request: Request):
    await get_current_user(request)
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.members.update_one({"id": member_id}, {"$set": update_data})
    updated = await db.members.find_one({"id": member_id}, {"_id": 0})
    return updated

@api_router.delete("/members/{member_id}")
async def delete_member(member_id: str, request: Request):
    await get_current_user(request)
    await db.members.delete_one({"id": member_id})
    return {"message": "Socio eliminado"}

# --- TRAINING SCHEDULES ENDPOINTS ---
@api_router.get("/training-schedules")
async def get_training_schedules(team_id: Optional[str] = None):
    query = {"team_id": team_id} if team_id else {}
    schedules = await db.training_schedules.find(query, {"_id": 0}).to_list(200)
    return schedules

@api_router.post("/training-schedules")
async def create_training_schedule(data: TrainingScheduleCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.training_schedules.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/training-schedules/{schedule_id}")
async def update_training_schedule(schedule_id: str, data: TrainingScheduleCreate, request: Request):
    await get_current_user(request)
    update_data = data.model_dump()
    await db.training_schedules.update_one({"id": schedule_id}, {"$set": update_data})
    updated = await db.training_schedules.find_one({"id": schedule_id}, {"_id": 0})
    return updated

@api_router.delete("/training-schedules/{schedule_id}")
async def delete_training_schedule(schedule_id: str, request: Request):
    await get_current_user(request)
    await db.training_schedules.delete_one({"id": schedule_id})
    return {"message": "Horario eliminado"}

# --- MATCHES ENDPOINTS ---
@api_router.get("/matches")
async def get_matches(status: Optional[str] = None):
    query = {"status": status} if status else {}
    matches = await db.matches.find(query, {"_id": 0}).sort("date", 1).to_list(200)
    return matches

@api_router.post("/matches")
async def create_match(data: MatchCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.matches.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/matches/{match_id}")
async def update_match(match_id: str, data: MatchCreate, request: Request):
    await get_current_user(request)
    update_data = data.model_dump()
    await db.matches.update_one({"id": match_id}, {"$set": update_data})
    updated = await db.matches.find_one({"id": match_id}, {"_id": 0})
    return updated

@api_router.delete("/matches/{match_id}")
async def delete_match(match_id: str, request: Request):
    await get_current_user(request)
    await db.matches.delete_one({"id": match_id})
    return {"message": "Partido eliminado"}

# --- GALLERY ENDPOINTS ---
@api_router.get("/gallery")
async def get_gallery(section: Optional[str] = None):
    query: dict = {}
    if section:
        query["section"] = section
    items = await db.gallery.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items

@api_router.get("/gallery/public")
async def get_gallery_public(section: Optional[str] = None):
    """Public endpoint — no auth required. Used by the HTML website to load images."""
    query: dict = {"visible": True}
    if section:
        query["section"] = section
    items = await db.gallery.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return items

@api_router.post("/gallery")
async def create_gallery_item(data: GalleryCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.gallery.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/gallery/{item_id}")
async def update_gallery_item(item_id: str, data: dict, request: Request):
    await get_current_user(request)
    await db.gallery.update_one({"id": item_id}, {"$set": data})
    updated = await db.gallery.find_one({"id": item_id}, {"_id": 0})
    return updated

@api_router.post("/gallery/upload")
async def upload_gallery_image(
    request: Request,
    file: UploadFile = File(...),
    title: str = Form(""),
    section: str = Form("galeria"),
    description: str = Form(""),
):
    """Upload an image file and create a gallery item."""
    await get_current_user(request)
    import os, base64
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Imagen demasiado grande (máx 10 MB)")
    ext = (file.filename or "img.jpg").rsplit(".", 1)[-1].lower()
    if ext not in ("jpg", "jpeg", "png", "gif", "webp", "svg"):
        raise HTTPException(status_code=400, detail="Formato no permitido")
    # Save to uploads directory
    upload_dir = os.path.join(os.path.dirname(__file__), "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    filename = f"{uuid.uuid4()}.{ext}"
    filepath = os.path.join(upload_dir, filename)
    with open(filepath, "wb") as f:
        f.write(content)
    image_url = f"/api/uploads/{filename}"
    doc = {
        "id": str(uuid.uuid4()),
        "title": title or file.filename,
        "image_url": image_url,
        "description": description,
        "section": section,
        "category": section,
        "visible": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.gallery.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.get("/uploads/{filename}")
async def serve_upload(filename: str):
    """Serve uploaded image files."""
    import os
    from fastapi.responses import FileResponse
    upload_dir = os.path.join(os.path.dirname(__file__), "uploads")
    filepath = os.path.join(upload_dir, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return FileResponse(filepath)

@api_router.delete("/gallery/{item_id}")
async def delete_gallery_item(item_id: str, request: Request):
    await get_current_user(request)
    item = await db.gallery.find_one({"id": item_id})
    if item and item.get("image_url", "").startswith("/api/uploads/"):
        import os
        filename = item["image_url"].split("/")[-1]
        filepath = os.path.join(os.path.dirname(__file__), "uploads", filename)
        if os.path.exists(filepath):
            os.remove(filepath)
    await db.gallery.delete_one({"id": item_id})
    return {"message": "Imagen eliminada"}

# --- CONTACT ENDPOINTS ---
@api_router.post("/contact")
async def submit_contact(data: ContactCreate):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["read"] = False
    await db.contacts.insert_one(doc)
    return {"message": "Mensaje enviado correctamente"}

@api_router.get("/contact")
async def get_contacts(request: Request):
    await get_current_user(request)
    contacts = await db.contacts.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return contacts

@api_router.put("/contact/{contact_id}/read")
async def mark_contact_read(contact_id: str, request: Request):
    await get_current_user(request)
    body = await request.json() if request.headers.get("content-length", "0") != "0" else {}
    read_value = body.get("read", True)
    await db.contacts.update_one({"id": contact_id}, {"$set": {"read": read_value}})
    return {"message": "Actualizado"}

@api_router.post("/contact/{contact_id}/reply")
async def reply_to_contact(contact_id: str, body: dict, request: Request):
    """Reply to a contact message using club SMTP settings."""
    await get_current_user(request)
    contact = await db.contacts.find_one({"id": contact_id})
    if not contact:
        raise HTTPException(status_code=404, detail="Mensaje no encontrado")
    reply_text = body.get("message", "")
    reply_subject = body.get("subject", f"Re: {contact.get('subject', 'Tu mensaje')}")
    to_email = contact.get("email", "")
    if not to_email:
        raise HTTPException(status_code=400, detail="El mensaje no tiene email de respuesta")
    settings = await db.settings.find_one({"type": "club"}, {"_id": 0}) or {}
    smtp_host = settings.get("smtp_host", "")
    smtp_port = int(settings.get("smtp_port", 587))
    smtp_user = settings.get("smtp_user", "")
    smtp_pass = settings.get("smtp_password", "")
    from_email = settings.get("email", smtp_user)
    if not smtp_host or not smtp_user:
        raise HTTPException(status_code=400, detail="SMTP no configurado. Ve a Ajustes → Email.")
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    msg = MIMEMultipart("alternative")
    msg["Subject"] = reply_subject
    msg["From"] = from_email
    msg["To"] = to_email
    html_body = f"""<div style="font-family:Arial,sans-serif;max-width:600px">
    <p>{reply_text.replace(chr(10), '<br>')}</p>
    <hr style="margin-top:32px;border:none;border-top:1px solid #eee">
    <p style="font-size:12px;color:#999">Mensaje original de {contact.get('name','')}: {contact.get('message','')}</p>
    </div>"""
    msg.attach(MIMEText(reply_text, "plain"))
    msg.attach(MIMEText(html_body, "html"))
    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(from_email, [to_email], msg.as_string())
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar: {str(e)}")
    await db.contacts.update_one({"id": contact_id}, {"$set": {"read": True, "replied": True, "replied_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Respuesta enviada correctamente"}

@api_router.delete("/contact/{contact_id}")
async def delete_contact(contact_id: str, request: Request):
    await get_current_user(request)
    await db.contacts.delete_one({"id": contact_id})
    return {"message": "Mensaje eliminado"}

# --- SETTINGS ENDPOINTS ---
@api_router.get("/settings")
async def get_settings():
    settings = await db.settings.find_one({"type": "club"}, {"_id": 0})
    if not settings:
        settings = {
            "type": "club",
            "club_name": "Racing San Gabriel A.D.C.",
            "description": "Escuela de futbol de referencia en Alicante. Ubicada en el corazon de Alacant, ofrecemos una experiencia deportiva de primer nivel. Futbol base, futbol femenino, futbol sala y mucho mas.",
            "address": "Carrer Racing San Gabriel, 39, 03008 Alacant, Alicante",
            "phone": "+34 617 50 27 80",
            "email": "racingsangabrieladc@hotmail.com",
            "instagram_url": "https://www.instagram.com/racingsangabrieladc/",
            "facebook_url": "https://www.facebook.com/RacingSanGabrielADC/",
            "instagram_username": "racingsangabrieladc",
            "facebook_page": "RacingSanGabrielADC",
            "facebook_embed_enabled": True,
            "facebook_page_url": "https://www.facebook.com/RacingSanGabrielADC/",
            "instagram_embed_code": "",
            "social_feed_provider": "none",
            "custom_embed_code": "",
            "schedule": "Lunes a Sabado: 9:00 - 21:00 | Domingo: Cerrado"
        }
        await db.settings.insert_one(settings)
        settings.pop("_id", None)
    return settings

@api_router.put("/settings")
async def update_settings(data: SettingsUpdate, request: Request):
    await get_current_user(request)
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.settings.update_one({"type": "club"}, {"$set": update_data}, upsert=True)
    settings = await db.settings.find_one({"type": "club"}, {"_id": 0})
    return settings

# --- FEES (TARIFAS) ENDPOINTS ---
@api_router.get("/fees")
async def get_fees(active_only: bool = False):
    query = {"active": True} if active_only else {}
    fees = await db.fees.find(query, {"_id": 0}).sort("name", 1).to_list(100)
    return fees

@api_router.post("/fees")
async def create_fee(data: FeeCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.fees.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/fees/{fee_id}")
async def update_fee(fee_id: str, data: FeeUpdate, request: Request):
    await get_current_user(request)
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.fees.update_one({"id": fee_id}, {"$set": update_data})
    updated = await db.fees.find_one({"id": fee_id}, {"_id": 0})
    return updated

@api_router.delete("/fees/{fee_id}")
async def delete_fee(fee_id: str, request: Request):
    await get_current_user(request)
    await db.fees.delete_one({"id": fee_id})
    return {"message": "Tarifa eliminada"}

# --- PAYMENTS (PAGOS) ENDPOINTS ---
@api_router.get("/payments")
async def get_payments(request: Request, person_id: Optional[str] = None, status: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if person_id:
        query["person_id"] = person_id
    if status:
        query["status"] = status
    payments = await db.payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return payments

@api_router.post("/payments")
async def create_payment(data: PaymentCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["stripe_session_id"] = ""
    doc["stripe_payment_intent_id"] = ""
    doc["paid_at"] = ""
    await db.payments.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/payments/{payment_id}")
async def update_payment(payment_id: str, data: PaymentUpdate, request: Request):
    await get_current_user(request)
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.payments.update_one({"id": payment_id}, {"$set": update_data})
    updated = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    return updated

@api_router.delete("/payments/{payment_id}")
async def delete_payment(payment_id: str, request: Request):
    await get_current_user(request)
    await db.payments.delete_one({"id": payment_id})
    return {"message": "Pago eliminado"}

@api_router.get("/payments/report/overdue")
async def get_overdue_report(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    overdue = await db.payments.find(
        {"status": "pending", "due_date": {"$lt": today, "$ne": ""}},
        {"_id": 0}
    ).sort("due_date", 1).to_list(500)
    return overdue

# --- PUBLIC REGISTRATION ENDPOINTS ---
@api_router.post("/register/player")
async def register_player(data: PlayerRegistrationRequest):
    """Public endpoint — creates pending player. Supports stripe, bank_transfer, redsys."""
    fee = await db.fees.find_one({"id": data.fee_id, "active": True}, {"_id": 0})
    if not fee:
        raise HTTPException(status_code=404, detail="Tarifa no encontrada")

    settings = await db.settings.find_one({"type": "club"}, {"_id": 0}) or {}
    club_name = settings.get("club_name", "Club Deportivo")

    # Detect if minor
    is_minor = False
    try:
        age = (datetime.now() - datetime.strptime(data.birthdate, "%Y-%m-%d")).days / 365.25
        is_minor = age < 18
    except Exception:
        pass

    if is_minor and not data.guardian:
        raise HTTPException(status_code=400, detail="Los jugadores menores de edad requieren datos del tutor legal")

    player_id = str(uuid.uuid4())
    family_id = str(uuid.uuid4())
    payment_method = data.payment_method or "stripe"

    # Validate payment method is enabled
    if payment_method == "stripe" and not settings.get("stripe_enabled"):
        stripe_key = get_stripe_key(settings)
        if not stripe_key:
            raise HTTPException(status_code=503, detail="El pago con tarjeta no está configurado. Contacta con el club.")
    if payment_method == "bank_transfer" and not settings.get("bank_transfer_enabled"):
        raise HTTPException(status_code=503, detail="La transferencia bancaria no está habilitada.")
    if payment_method == "redsys" and not settings.get("redsys_enabled"):
        raise HTTPException(status_code=503, detail="El TPV Redsys no está configurado.")

    # Build player document (common for all methods)
    player_doc = {
        "id": player_id,
        "name": data.name,
        "surname": data.surname,
        "birthdate": data.birthdate,
        "dni": data.dni or "",
        "phone": data.phone,
        "email": data.email,
        "address": data.address or "",
        "city": data.city or "",
        "postal_code": data.postal_code or "",
        "team_id": data.team_id,
        "position": data.position or "",
        "jersey_size": data.jersey_size or "",
        "medical_notes": data.medical_notes or "",
        "family_id": family_id,
        "status": "pending_payment",
        "season": "2025/2026",
        "notes": "",
        "image_url": "",
        "number": None,
        "blood_type": "",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "stripe_session_id": "",
        "payment_method": payment_method,
        # RGPD — consentimiento registrado
        "consent_gdpr": data.consent_gdpr,
        "consent_communications": data.consent_communications,
        "consent_guardian_gdpr": data.consent_guardian_gdpr if is_minor else None,
        "consent_date": datetime.now(timezone.utc).isoformat(),
        "consent_ip": request.client.host if request.client else "",
    }

    payment_doc = {
        "id": str(uuid.uuid4()),
        "person_id": player_id,
        "person_type": "player",
        "fee_id": data.fee_id,
        "amount": fee["amount"],
        "concept": fee["name"],
        "status": "pending",
        "stripe_session_id": "",
        "stripe_payment_intent_id": "",
        "due_date": "",
        "paid_at": "",
        "notes": f"Pago via {payment_method}",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    # --- STRIPE ---
    if payment_method == "stripe":
        stripe_key = get_stripe_key(settings)
        if not stripe_key:
            raise HTTPException(status_code=503, detail="Stripe no configurado.")
        stripe.api_key = stripe_key
        success_url = data.success_url or f"{os.environ.get('FRONTEND_URL','http://localhost:3000')}/pago/exito?session_id={{CHECKOUT_SESSION_ID}}"
        cancel_url = data.cancel_url or f"{os.environ.get('FRONTEND_URL','http://localhost:3000')}/pago/cancelado"
        try:
            session = stripe.checkout.Session.create(
                payment_method_types=["card"],
                line_items=[{
                    "price_data": {
                        "currency": fee.get("currency", "eur"),
                        "product_data": {
                            "name": f"{fee['name']} — {data.name} {data.surname}",
                            "description": f"{club_name} · {fee.get('description','')}",
                        },
                        "unit_amount": int(fee["amount"] * 100),
                    },
                    "quantity": 1,
                }],
                mode="payment",
                success_url=success_url,
                cancel_url=cancel_url,
                customer_email=data.email,
                metadata={"type": "player_registration", "player_id": player_id, "family_id": family_id, "fee_id": data.fee_id},
            )
        except stripe.StripeError as e:
            raise HTTPException(status_code=400, detail=str(e))
        player_doc["stripe_session_id"] = session.id
        payment_doc["stripe_session_id"] = session.id
        payment_doc["notes"] = "Pago online via Stripe"
        await db.players.insert_one(player_doc)
        await _store_guardian_and_payment(is_minor, data, player_id, family_id, payment_doc)
        await send_registration_confirmation("player", player_doc, settings)
        return {"payment_method": "stripe", "checkout_url": session.url, "session_id": session.id}

    # --- BANK TRANSFER ---
    if payment_method == "bank_transfer":
        player_doc["status"] = "pending_payment"
        payment_doc["notes"] = "Pendiente transferencia bancaria"
        await db.players.insert_one(player_doc)
        await _store_guardian_and_payment(is_minor, data, player_id, family_id, payment_doc)
        await send_registration_confirmation("player", player_doc, settings)
        return {
            "payment_method": "bank_transfer",
            "player_id": player_id,
            "bank_iban": settings.get("bank_iban", ""),
            "bank_bic": settings.get("bank_bic", ""),
            "bank_name": settings.get("bank_name", ""),
            "bank_holder": settings.get("bank_holder", ""),
            "bank_transfer_instructions": settings.get("bank_transfer_instructions", ""),
            "amount": fee["amount"],
            "concept": f"Inscripción {data.name} {data.surname}",
        }

    # --- REDSYS ---
    if payment_method == "redsys":
        if not settings.get("redsys_secret_key"):
            raise HTTPException(status_code=503, detail="Redsys no configurado correctamente.")
        await db.players.insert_one(player_doc)
        await _store_guardian_and_payment(is_minor, data, player_id, family_id, payment_doc)
        order = player_id[:12].replace("-", "").upper()[:12].zfill(4)
        params = {
            "DS_MERCHANT_AMOUNT": str(int(fee["amount"] * 100)),
            "DS_MERCHANT_ORDER": order,
            "DS_MERCHANT_MERCHANTCODE": settings["redsys_merchant_code"],
            "DS_MERCHANT_CURRENCY": "978",
            "DS_MERCHANT_TRANSACTIONTYPE": "0",
            "DS_MERCHANT_TERMINAL": settings.get("redsys_terminal", "001"),
            "DS_MERCHANT_MERCHANTURL": f"{os.environ.get('BACKEND_URL','')}/api/redsys/notification",
            "DS_MERCHANT_URLOK": f"{os.environ.get('FRONTEND_URL','http://localhost:3000')}/pago/exito",
            "DS_MERCHANT_URLKO": f"{os.environ.get('FRONTEND_URL','http://localhost:3000')}/pago/cancelado",
            "DS_MERCHANT_PRODUCTDESCRIPTION": f"{fee['name']} - {data.name} {data.surname}"[:125],
            "DS_MERCHANT_MERCHANTDATA": payment_doc["id"],
        }
        encoded = redsys_encode_params(params)
        try:
            signature = redsys_sign(encoded, settings["redsys_secret_key"], order)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error Redsys: {e}")
        env = settings.get("redsys_environment", "test")
        gateway_url = "https://sis-t.redsys.es:25443/sis/realizarPago" if env == "test" else "https://sis.redsys.es/sis/realizarPago"
        return {
            "payment_method": "redsys",
            "gateway_url": gateway_url,
            "Ds_SignatureVersion": "HMAC_SHA256_V1",
            "Ds_MerchantParameters": encoded,
            "Ds_Signature": signature,
        }

    raise HTTPException(status_code=400, detail="Método de pago no soportado")

async def _store_guardian_and_payment(is_minor: bool, data, player_id: str, family_id: str, payment_doc: dict):
    """Helper to store guardian (if minor) and payment record."""
    if is_minor and data.guardian:
        guardian_doc = {
            "id": str(uuid.uuid4()),
            "name": data.guardian.name,
            "surname": data.guardian.surname or "",
            "dni": data.guardian.dni or "",
            "phone": data.guardian.phone,
            "email": data.guardian.email,
            "relationship": data.guardian.relationship or "padre",
            "player_ids": [player_id],
            "family_id": family_id,
            "bank_iban": "",
            "address": "",
            "city": "",
            "notes": "Registrado via formulario web",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.guardians.insert_one(guardian_doc)
    await db.payments.insert_one(payment_doc)

@api_router.post("/register/member")
async def register_member(data: MemberRegistrationRequest, request: Request):
    """Public endpoint — creates pending member. Supports stripe, bank_transfer, redsys."""
    fee = await db.fees.find_one({"id": data.fee_id, "active": True}, {"_id": 0})
    if not fee:
        raise HTTPException(status_code=404, detail="Tarifa no encontrada")

    settings = await db.settings.find_one({"type": "club"}, {"_id": 0}) or {}
    club_name = settings.get("club_name", "Club Deportivo")

    member_id = str(uuid.uuid4())
    count = await db.members.count_documents({})
    member_number = f"S{count + 1:04d}"
    payment_method = data.payment_method or "stripe"

    member_doc = {
        "id": member_id,
        "name": data.name,
        "surname": data.surname,
        "birthdate": data.birthdate or "",
        "dni": data.dni or "",
        "phone": data.phone,
        "email": data.email,
        "address": data.address or "",
        "city": data.city or "",
        "postal_code": data.postal_code or "",
        "member_type": data.member_type or "socio_adulto",
        "bank_iban": encrypt_iban(data.bank_iban) if data.bank_iban else "",
        "photo_url": "",
        "status": "pending_payment",
        "season": "2025/2026",
        "notes": "Registrado via formulario web",
        "member_number": member_number,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "stripe_session_id": "",
        "payment_method": payment_method,
        # RGPD — consentimiento registrado
        "consent_gdpr": data.consent_gdpr,
        "consent_communications": data.consent_communications,
        "consent_date": datetime.now(timezone.utc).isoformat(),
        "consent_ip": request.client.host if request.client else "",
    }

    payment_doc = {
        "id": str(uuid.uuid4()),
        "person_id": member_id,
        "person_type": "member",
        "fee_id": data.fee_id,
        "amount": fee["amount"],
        "concept": fee["name"],
        "status": "pending",
        "stripe_session_id": "",
        "stripe_payment_intent_id": "",
        "due_date": "",
        "paid_at": "",
        "notes": f"Pago via {payment_method}",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    # --- STRIPE ---
    if payment_method == "stripe":
        stripe_key = get_stripe_key(settings)
        if not stripe_key:
            raise HTTPException(status_code=503, detail="Stripe no configurado.")
        stripe.api_key = stripe_key
        success_url = data.success_url or f"{os.environ.get('FRONTEND_URL','http://localhost:3000')}/pago/exito?session_id={{CHECKOUT_SESSION_ID}}"
        cancel_url = data.cancel_url or f"{os.environ.get('FRONTEND_URL','http://localhost:3000')}/pago/cancelado"
        try:
            session = stripe.checkout.Session.create(
                payment_method_types=["card"],
                line_items=[{
                    "price_data": {
                        "currency": fee.get("currency", "eur"),
                        "product_data": {
                            "name": f"{fee['name']} — {data.name} {data.surname}",
                            "description": f"{club_name}",
                        },
                        "unit_amount": int(fee["amount"] * 100),
                    },
                    "quantity": 1,
                }],
                mode="payment",
                success_url=success_url,
                cancel_url=cancel_url,
                customer_email=data.email,
                metadata={"type": "member_registration", "member_id": member_id, "fee_id": data.fee_id},
            )
        except stripe.StripeError as e:
            raise HTTPException(status_code=400, detail=str(e))
        member_doc["stripe_session_id"] = session.id
        payment_doc["stripe_session_id"] = session.id
        payment_doc["notes"] = "Pago online via Stripe"
        await db.members.insert_one(member_doc)
        await db.payments.insert_one(payment_doc)
        await send_registration_confirmation("member", member_doc, settings)
        return {"payment_method": "stripe", "checkout_url": session.url, "session_id": session.id}

    # --- BANK TRANSFER ---
    if payment_method == "bank_transfer":
        if not settings.get("bank_transfer_enabled"):
            raise HTTPException(status_code=503, detail="Transferencia bancaria no habilitada.")
        payment_doc["notes"] = "Pendiente transferencia bancaria"
        await db.members.insert_one(member_doc)
        await db.payments.insert_one(payment_doc)
        await send_registration_confirmation("member", member_doc, settings)
        return {
            "payment_method": "bank_transfer",
            "member_id": member_id,
            "member_number": member_number,
            "bank_iban": settings.get("bank_iban", ""),
            "bank_bic": settings.get("bank_bic", ""),
            "bank_name": settings.get("bank_name", ""),
            "bank_holder": settings.get("bank_holder", ""),
            "bank_transfer_instructions": settings.get("bank_transfer_instructions", ""),
            "amount": fee["amount"],
            "concept": f"Alta socio {data.name} {data.surname} — {member_number}",
        }

    # --- REDSYS ---
    if payment_method == "redsys":
        if not settings.get("redsys_enabled") or not settings.get("redsys_secret_key"):
            raise HTTPException(status_code=503, detail="Redsys no configurado.")
        await db.members.insert_one(member_doc)
        await db.payments.insert_one(payment_doc)
        order = member_id[:12].replace("-", "").upper()[:12].zfill(4)
        params = {
            "DS_MERCHANT_AMOUNT": str(int(fee["amount"] * 100)),
            "DS_MERCHANT_ORDER": order,
            "DS_MERCHANT_MERCHANTCODE": settings["redsys_merchant_code"],
            "DS_MERCHANT_CURRENCY": "978",
            "DS_MERCHANT_TRANSACTIONTYPE": "0",
            "DS_MERCHANT_TERMINAL": settings.get("redsys_terminal", "001"),
            "DS_MERCHANT_MERCHANTURL": f"{os.environ.get('BACKEND_URL','')}/api/redsys/notification",
            "DS_MERCHANT_URLOK": f"{os.environ.get('FRONTEND_URL','http://localhost:3000')}/pago/exito",
            "DS_MERCHANT_URLKO": f"{os.environ.get('FRONTEND_URL','http://localhost:3000')}/pago/cancelado",
            "DS_MERCHANT_PRODUCTDESCRIPTION": f"{fee['name']} - {data.name} {data.surname}"[:125],
            "DS_MERCHANT_MERCHANTDATA": payment_doc["id"],
        }
        encoded = redsys_encode_params(params)
        try:
            signature = redsys_sign(encoded, settings["redsys_secret_key"], order)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error Redsys: {e}")
        env = settings.get("redsys_environment", "test")
        gateway_url = "https://sis-t.redsys.es:25443/sis/realizarPago" if env == "test" else "https://sis.redsys.es/sis/realizarPago"
        return {
            "payment_method": "redsys",
            "gateway_url": gateway_url,
            "Ds_SignatureVersion": "HMAC_SHA256_V1",
            "Ds_MerchantParameters": encoded,
            "Ds_Signature": signature,
        }

    raise HTTPException(status_code=400, detail="Método de pago no soportado")

# --- STRIPE WEBHOOK ---
@api_router.post("/stripe/webhook")
async def stripe_webhook(request: Request):
    """Stripe calls this after a successful payment."""
    payload = await request.body()
    sig_header = request.headers.get("stripe-signature", "")
    webhook_secret = os.environ.get("STRIPE_WEBHOOK_SECRET", "")

    try:
        if webhook_secret:
            event = stripe.Webhook.construct_event(payload, sig_header, webhook_secret)
        else:
            event = stripe.Event.construct_from({"type": "checkout.session.completed", "data": {"object": {}}}, stripe.api_key)
            import json
            event = stripe.Event.construct_from(json.loads(payload), stripe.api_key)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    if event["type"] == "checkout.session.completed":
        session_obj = event["data"]["object"]
        session_id = session_obj.get("id")
        payment_intent_id = session_obj.get("payment_intent", "")
        meta = session_obj.get("metadata", {})
        reg_type = meta.get("type")
        paid_at = datetime.now(timezone.utc).isoformat()

        if reg_type == "player_registration":
            player_id = meta.get("player_id")
            await db.players.update_one(
                {"id": player_id},
                {"$set": {"status": "active"}}
            )
        elif reg_type == "member_registration":
            member_id = meta.get("member_id")
            await db.members.update_one(
                {"id": member_id},
                {"$set": {"status": "active"}}
            )

        await db.payments.update_one(
            {"stripe_session_id": session_id},
            {"$set": {"status": "paid", "paid_at": paid_at, "stripe_payment_intent_id": payment_intent_id}}
        )

    return {"received": True}

# --- EMAIL UTILITY ---
async def send_email(settings: dict, to_email: str, subject: str, body_html: str, body_text: str = ""):
    """Send email using the club's SMTP credentials stored in settings."""
    if not settings.get("smtp_enabled") or not settings.get("smtp_host"):
        return False
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = f"{settings.get('smtp_from_name', settings.get('club_name', 'Club'))} <{settings['smtp_user']}>"
        msg["To"] = to_email
        if body_text:
            msg.attach(MIMEText(body_text, "plain", "utf-8"))
        msg.attach(MIMEText(body_html, "html", "utf-8"))
        port = int(settings.get("smtp_port", 587))
        use_tls = settings.get("smtp_use_tls", True)
        with smtplib.SMTP(settings["smtp_host"], port, timeout=10) as server:
            if use_tls:
                server.starttls()
            server.login(settings["smtp_user"], settings["smtp_password"])
            server.sendmail(settings["smtp_user"], to_email, msg.as_string())
        return True
    except Exception as e:
        logger.error(f"Email error: {e}")
        return False

async def get_club_settings():
    s = await db.settings.find_one({"type": "club"}, {"_id": 0})
    return s or {}

def get_stripe_key(settings: dict) -> str:
    """Returns club's Stripe key if configured, else falls back to env var."""
    return settings.get("stripe_secret_key") or os.environ.get("STRIPE_SECRET_KEY", "")

# --- EMAIL TEST ENDPOINT ---
@api_router.post("/settings/test-email")
async def test_email(request: Request):
    user = await get_current_user(request)
    settings = await get_club_settings()
    ok = await send_email(
        settings, user["email"],
        "✅ Prueba de email configurado correctamente",
        f"<p>El correo del club <b>{settings.get('club_name','')}</b> está configurado correctamente.</p>",
    )
    if not ok:
        raise HTTPException(status_code=400, detail="No se pudo enviar el email. Revisa la configuración SMTP.")
    return {"message": f"Email de prueba enviado a {user['email']}"}

# --- REDSYS UTILITIES ---
def redsys_encode_params(params: dict) -> str:
    return base64.b64encode(json.dumps(params).encode()).decode()

def redsys_sign(encoded_params: str, secret_key: str, order: str) -> str:
    key = base64.b64decode(secret_key)
    # Ensure 24-byte key for 3DES
    if len(key) == 16:
        key = key + key[:8]
    # PKCS7-pad the order to a multiple of 8 bytes, then 3DES-CBC with zero IV
    order_bytes = order.encode("ascii")
    padder = sym_padding.PKCS7(64).padder()
    padded_order = padder.update(order_bytes) + padder.finalize()
    cipher = Cipher(algorithms.TripleDES(key), modes.CBC(b"\x00" * 8), backend=default_backend())
    encryptor = cipher.encryptor()
    diversified = encryptor.update(padded_order) + encryptor.finalize()
    # HMAC-SHA256 of the encoded params using the diversified key
    sig = hmac_lib.new(diversified, encoded_params.encode("utf-8"), hashlib.sha256).digest()
    return base64.b64encode(sig).decode()

@api_router.post("/redsys/create-payment")
async def redsys_create_payment(request: Request, payment_id: str, amount: float, order_ref: str, concept: str):
    settings = await get_club_settings()
    if not settings.get("redsys_enabled") or not settings.get("redsys_merchant_code"):
        raise HTTPException(status_code=503, detail="Redsys no está configurado")
    merchant_code = settings["redsys_merchant_code"]
    terminal = settings.get("redsys_terminal", "001")
    secret_key = settings["redsys_secret_key"]
    env = settings.get("redsys_environment", "test")
    frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
    order = order_ref[:12].zfill(4)
    params = {
        "DS_MERCHANT_AMOUNT": str(int(amount * 100)),
        "DS_MERCHANT_ORDER": order,
        "DS_MERCHANT_MERCHANTCODE": merchant_code,
        "DS_MERCHANT_CURRENCY": "978",
        "DS_MERCHANT_TRANSACTIONTYPE": "0",
        "DS_MERCHANT_TERMINAL": terminal,
        "DS_MERCHANT_MERCHANTURL": f"{os.environ.get('BACKEND_URL','')}/api/redsys/notification",
        "DS_MERCHANT_URLOK": f"{frontend_url}/pago/exito",
        "DS_MERCHANT_URLKO": f"{frontend_url}/pago/cancelado",
        "DS_MERCHANT_PRODUCTDESCRIPTION": concept[:125],
        "DS_MERCHANT_CONSUMERLANGUAGE": "003",
    }
    encoded = redsys_encode_params(params)
    try:
        signature = redsys_sign(encoded, secret_key, order)
    except Exception:
        raise HTTPException(status_code=500, detail="Error al firmar la petición Redsys. Verifica la clave secreta.")
    gateway_url = (
        "https://sis-t.redsys.es:25443/sis/realizarPago"
        if env == "test"
        else "https://sis.redsys.es/sis/realizarPago"
    )
    return {
        "gateway_url": gateway_url,
        "Ds_SignatureVersion": "HMAC_SHA256_V1",
        "Ds_MerchantParameters": encoded,
        "Ds_Signature": signature,
    }

@api_router.post("/redsys/notification")
async def redsys_notification(request: Request):
    """Redsys calls this URL to notify payment result."""
    form = await request.form()
    encoded_params = form.get("Ds_MerchantParameters", "")
    signature = form.get("Ds_Signature", "")
    settings = await get_club_settings()
    if not settings.get("redsys_secret_key"):
        raise HTTPException(status_code=400, detail="Redsys no configurado")
    try:
        params = json.loads(base64.b64decode(encoded_params).decode())
        order = params.get("Ds_Order", "")
        expected_sig = redsys_sign(encoded_params, settings["redsys_secret_key"], order)
        if expected_sig != signature:
            raise HTTPException(status_code=400, detail="Firma inválida")
        response_code = int(params.get("Ds_Response", "9999"))
        payment_id = params.get("Ds_MerchantData", "")
        if response_code < 100:
            await db.payments.update_one(
                {"id": payment_id},
                {"$set": {"status": "paid", "paid_at": datetime.now(timezone.utc).isoformat()}}
            )
    except Exception as e:
        logger.error(f"Redsys notification error: {e}")
    return {"status": "ok"}

# --- SEPA HELPERS ---
_SEPA_ALLOWED = set("ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789 /-?:().,'+")

def sepa_str(text: str, max_len: int = 70) -> str:
    """Normalize a string to the SEPA Latin character set and truncate.
    EPC SEPA rulebook allows: A-Z a-z 0-9 / - ? : ( ) . , ' + SPACE.
    Accented chars are converted to their ASCII base; others are dropped."""
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    base = "".join(c for c in nfkd if not unicodedata.combining(c))
    clean = "".join(c if c in _SEPA_ALLOWED else " " for c in base)
    clean = " ".join(clean.split())  # collapse whitespace
    return clean[:max_len]

def sepa_iban(iban: str) -> str:
    return iban.replace(" ", "").upper()

# --- CIFRADO DE IBAN (RGPD Art. 32) ---
from cryptography.fernet import Fernet as _Fernet

def _fernet():
    key_raw = os.environ.get("SECRET_KEY", "sudeporte-default-key-CHANGE-IN-PROD")
    key_bytes = hashlib.sha256(key_raw.encode()).digest()
    return _Fernet(base64.urlsafe_b64encode(key_bytes))

def encrypt_iban(iban: str) -> str:
    """Cifra un IBAN para almacenamiento seguro. Devuelve 'ENC:...'"""
    if not iban or iban.upper().startswith("ENC:"):
        return iban
    clean = iban.replace(" ", "").upper()
    try:
        return "ENC:" + _fernet().encrypt(clean.encode()).decode()
    except Exception:
        return iban

def decrypt_iban(iban: str) -> str:
    """Descifra un IBAN almacenado. Compatible con IBANs no cifrados (migraciones)."""
    if not iban or not iban.upper().startswith("ENC:"):
        return iban
    try:
        return _fernet().decrypt(iban[4:].encode()).decode()
    except Exception:
        return iban  # fallback: devuelve como está (datos previos sin cifrar)

def mask_iban(iban: str) -> str:
    """Devuelve IBAN enmascarado para mostrar en pantalla: ES76 •••• •••• 6789"""
    plain = decrypt_iban(iban)
    clean = plain.replace(" ", "")
    if len(clean) < 8:
        return clean
    return f"{clean[:4]} •••• •••• {clean[-4:]}"

# --- SEPA MANDATES ---
@api_router.get("/sepa/mandates")
async def get_sepa_mandates(request: Request, person_id: Optional[str] = None):
    await get_current_user(request)
    query = {"person_id": person_id} if person_id else {}
    mandates = await db.sepa_mandates.find(query, {"_id": 0}).sort("signature_date", -1).to_list(500)
    return mandates

@api_router.post("/sepa/mandates")
async def create_sepa_mandate(data: SepaMandateCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["mandate_ref"] = data.mandate_ref or f"MND-{str(uuid.uuid4())[:8].upper()}"
    doc["debtor_iban"] = encrypt_iban(data.debtor_iban)  # cifrar IBAN en reposo
    doc["status"] = "active"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.sepa_mandates.insert_one(doc)
    result = {k: v for k, v in doc.items() if k != "_id"}
    result["debtor_iban"] = mask_iban(doc["debtor_iban"])  # devolver enmascarado
    return result

@api_router.delete("/sepa/mandates/{mandate_id}")
async def delete_sepa_mandate(mandate_id: str, request: Request):
    await get_current_user(request)
    await db.sepa_mandates.update_one({"id": mandate_id}, {"$set": {"status": "cancelled"}})
    return {"message": "Mandato cancelado"}

@api_router.post("/sepa/generate-xml")
async def generate_sepa_xml(data: SepaXmlRequest, request: Request):
    """
    Genera XML pain.008.001.02 conforme al esquema EPC SEPA Core Direct Debit
    (EPC130-08) y al Reglamento UE 260/2012.
    - FRST/OOFF y RCUR/FNAL van en bloques PmtInf separados (exigido por los bancos).
    - Todos los textos se normalizan al juego de caracteres SEPA (ASCII latino).
    - Límites de campo respetados: EndToEndId/MndtId ≤35, Nm ≤70, Ustrd ≤140.
    """
    await get_current_user(request)
    settings = await get_club_settings()

    creditor_id   = settings.get("sepa_creditor_id", "").replace(" ", "")
    creditor_name = sepa_str(settings.get("sepa_creditor_name") or settings.get("club_name", "Club"), 70)
    creditor_iban = sepa_iban(settings.get("sepa_creditor_iban", ""))
    creditor_bic  = (settings.get("sepa_creditor_bic") or "NOTPROVIDED").replace(" ", "").upper()

    if not creditor_id or not creditor_iban:
        raise HTTPException(status_code=400,
            detail="Configura el Identificador de Acreedor SEPA (ICS) y el IBAN del club en Ajustes → SEPA.")

    mandates = await db.sepa_mandates.find(
        {"id": {"$in": data.mandate_ids}, "status": "active"}, {"_id": 0}
    ).to_list(500)
    if not mandates:
        raise HTTPException(status_code=400, detail="No se encontraron mandatos activos con los IDs indicados.")

    # Resolver importe: 1) importe del mandato  2) default_amount del request  3) pago pendiente en BD
    tx_data = []
    for m in mandates:
        if m.get("amount") and float(m["amount"]) > 0:
            amt, payment = float(m["amount"]), None
        elif data.default_amount and float(data.default_amount) > 0:
            amt, payment = float(data.default_amount), None
        else:
            payment = await db.payments.find_one(
                {"person_id": m["person_id"], "status": "pending"},
                sort=[("created_at", -1)], projection={"_id": 0}
            )
            amt = float(payment["amount"]) if payment and payment.get("amount") else 0.0
        if amt > 0:
            tx_data.append({"mandate": m, "amount": round(amt, 2), "payment": payment})

    if not tx_data:
        raise HTTPException(status_code=400,
            detail="Ningún mandato tiene importe. Indica un importe por defecto en el diálogo o asigna cuota en cada mandato.")

    now         = datetime.now(timezone.utc)
    remesa_num  = data.remesa_number or int(now.strftime("%Y%m%d%H%M"))
    rand_hex    = uuid.uuid4().hex[:22]
    # MsgId ≤35 chars: "NNNNN-ID-xxxxxxxx" (5+3+22 = 30 chars max)
    msg_id      = f"{remesa_num}-ID-{rand_hex}"[:35]
    total       = round(sum(t["amount"] for t in tx_data), 2)

    # Separar FRST/OOFF (presentación inicial) de RCUR/FNAL (recurrentes)
    first_group = [t for t in tx_data if t["mandate"].get("sequence_type") in ("FRST", "OOFF")]
    recur_group = [t for t in tx_data if t["mandate"].get("sequence_type") not in ("FRST", "OOFF")]

    ns  = "urn:iso:std:iso:20022:tech:xsd:pain.008.001.02"
    xsi = "http://www.w3.org/2001/XMLSchema-instance"
    ET.register_namespace("", ns)
    ET.register_namespace("xsi", xsi)

    doc_el = ET.Element(f"{{{ns}}}Document")
    doc_el.set(f"{{{xsi}}}schemaLocation", f"{ns} pain.008.001.02.xsd")
    root = ET.SubElement(doc_el, f"{{{ns}}}CstmrDrctDbtInitn")

    # ── Cabecera de grupo (GrpHdr) ──────────────────────────────────────────
    grp = ET.SubElement(root, f"{{{ns}}}GrpHdr")
    ET.SubElement(grp, f"{{{ns}}}MsgId").text    = msg_id
    ET.SubElement(grp, f"{{{ns}}}CreDtTm").text  = now.strftime("%Y-%m-%dT%H:%M:%S+00:00")
    ET.SubElement(grp, f"{{{ns}}}NbOfTxs").text  = str(len(tx_data))
    ET.SubElement(grp, f"{{{ns}}}CtrlSum").text  = f"{total:.2f}"
    ip = ET.SubElement(grp, f"{{{ns}}}InitgPty")
    ET.SubElement(ip, f"{{{ns}}}Nm").text = creditor_name
    # El ICS también va en InitgPty/Id según las guías de implementación españolas
    ip_othr = ET.SubElement(ET.SubElement(ET.SubElement(ip, f"{{{ns}}}Id"), f"{{{ns}}}OrgId"), f"{{{ns}}}Othr")
    ET.SubElement(ip_othr, f"{{{ns}}}Id").text = creditor_id

    def add_pmt_inf(block_num: int, group: list, seq_tp: str):
        block_total = round(sum(t["amount"] for t in group), 2)
        pmt = ET.SubElement(root, f"{{{ns}}}PmtInf")
        # PmtInfId ≤35
        ET.SubElement(pmt, f"{{{ns}}}PmtInfId").text = f"{msg_id}/{block_num}"[:35]
        ET.SubElement(pmt, f"{{{ns}}}PmtMtd").text   = "DD"
        ET.SubElement(pmt, f"{{{ns}}}BtchBookg").text = "false"
        ET.SubElement(pmt, f"{{{ns}}}NbOfTxs").text  = str(len(group))
        ET.SubElement(pmt, f"{{{ns}}}CtrlSum").text  = f"{block_total:.2f}"
        # PmtTpInf
        pti = ET.SubElement(pmt, f"{{{ns}}}PmtTpInf")
        ET.SubElement(ET.SubElement(pti, f"{{{ns}}}SvcLvl"),    f"{{{ns}}}Cd").text = "SEPA"
        ET.SubElement(ET.SubElement(pti, f"{{{ns}}}LclInstrm"), f"{{{ns}}}Cd").text = "CORE"
        ET.SubElement(pti, f"{{{ns}}}SeqTp").text = seq_tp
        ET.SubElement(pmt, f"{{{ns}}}ReqdColltnDt").text = data.collection_date
        # Acreedor
        ET.SubElement(ET.SubElement(pmt, f"{{{ns}}}Cdtr"), f"{{{ns}}}Nm").text = creditor_name
        ET.SubElement(
            ET.SubElement(ET.SubElement(pmt, f"{{{ns}}}CdtrAcct"), f"{{{ns}}}Id"),
            f"{{{ns}}}IBAN"
        ).text = creditor_iban
        # CdtrAgt — BIC obligatorio si está disponible
        fii = ET.SubElement(ET.SubElement(pmt, f"{{{ns}}}CdtrAgt"), f"{{{ns}}}FinInstnId")
        if creditor_bic and creditor_bic != "NOTPROVIDED":
            ET.SubElement(fii, f"{{{ns}}}BIC").text = creditor_bic
        else:
            ET.SubElement(ET.SubElement(fii, f"{{{ns}}}Othr"), f"{{{ns}}}Id").text = "NOTPROVIDED"
        ET.SubElement(pmt, f"{{{ns}}}ChrgBr").text = "SLEV"
        # CdtrSchmeId — ICS del acreedor
        csi_othr = ET.SubElement(
            ET.SubElement(ET.SubElement(ET.SubElement(pmt, f"{{{ns}}}CdtrSchmeId"), f"{{{ns}}}Id"), f"{{{ns}}}PrvtId"),
            f"{{{ns}}}Othr"
        )
        ET.SubElement(csi_othr, f"{{{ns}}}Id").text = creditor_id
        ET.SubElement(ET.SubElement(csi_othr, f"{{{ns}}}SchmeNm"), f"{{{ns}}}Prtry").text = "SEPA"

        # ── Transacciones ─────────────────────────────────────────────────
        for idx, tx in enumerate(group):
            m   = tx["mandate"]
            amt = tx["amount"]
            # EndToEndId ≤35: usamos contador global de la remesa
            e2e_id = f"{remesa_num}-{(block_num - 1) * 10000 + idx + 1}"[:35]
            # MndtId ≤35
            mndt_id = sepa_str(m["mandate_ref"], 35)
            # Nombre deudor ≤70, charset SEPA
            dbtr_nm = sepa_str(m["person_name"], 70)

            ddi = ET.SubElement(pmt, f"{{{ns}}}DrctDbtTxInf")
            ET.SubElement(ET.SubElement(ddi, f"{{{ns}}}PmtId"), f"{{{ns}}}EndToEndId").text = e2e_id
            ia = ET.SubElement(ddi, f"{{{ns}}}InstdAmt")
            ia.set("Ccy", "EUR")
            ia.text = f"{amt:.2f}"
            # Información del mandato
            mri = ET.SubElement(ET.SubElement(ddi, f"{{{ns}}}DrctDbtTx"), f"{{{ns}}}MndtRltdInf")
            ET.SubElement(mri, f"{{{ns}}}MndtId").text     = mndt_id
            ET.SubElement(mri, f"{{{ns}}}DtOfSgntr").text  = m["signature_date"]
            # Banco deudor
            dfii = ET.SubElement(ET.SubElement(ddi, f"{{{ns}}}DbtrAgt"), f"{{{ns}}}FinInstnId")
            dbic = (m.get("debtor_bic") or "").replace(" ", "").upper()
            if dbic:
                ET.SubElement(dfii, f"{{{ns}}}BIC").text = dbic
            else:
                ET.SubElement(ET.SubElement(dfii, f"{{{ns}}}Othr"), f"{{{ns}}}Id").text = "NOTPROVIDED"
            # Deudor
            ET.SubElement(ET.SubElement(ddi, f"{{{ns}}}Dbtr"), f"{{{ns}}}Nm").text = dbtr_nm
            ET.SubElement(
                ET.SubElement(ET.SubElement(ddi, f"{{{ns}}}DbtrAcct"), f"{{{ns}}}Id"),
                f"{{{ns}}}IBAN"
            ).text = sepa_iban(decrypt_iban(m["debtor_iban"]))
            # Concepto
            ET.SubElement(ET.SubElement(ddi, f"{{{ns}}}RmtInf"), f"{{{ns}}}Ustrd").text = sepa_str(data.concept, 140)

    block = 1
    if first_group:
        add_pmt_inf(block, first_group, "FRST")
        block += 1
    if recur_group:
        add_pmt_inf(block, recur_group, "RCUR")

    # ── Post-generación ─────────────────────────────────────────────────────
    for tx in tx_data:
        m = tx["mandate"]
        # FRST pasa a RCUR para la siguiente remesa
        if m.get("sequence_type") == "FRST":
            await db.sepa_mandates.update_one({"id": m["id"]}, {"$set": {"sequence_type": "RCUR"}})
        # Marcar pago como cobrado
        if tx.get("payment"):
            await db.payments.update_one(
                {"id": tx["payment"]["id"]},
                {"$set": {"status": "paid", "paid_at": now.isoformat(), "notes": f"SEPA remesa {remesa_num}"}}
            )
    # Anotar nº de remesa y fecha en el mandato para trazabilidad
    mandate_ids_used = [t["mandate"]["id"] for t in tx_data]
    await db.sepa_mandates.update_many(
        {"id": {"$in": mandate_ids_used}},
        {"$set": {"last_remesa": remesa_num, "last_collection_date": data.collection_date}}
    )

    # ── Serializar XML ──────────────────────────────────────────────────────
    xml_bytes = io.BytesIO()
    tree = ET.ElementTree(doc_el)
    ET.indent(tree, space="  ")
    tree.write(xml_bytes, encoding="UTF-8", xml_declaration=True)
    xml_bytes.seek(0)

    filename = f"SepaCore_{remesa_num}_{data.collection_date}.xml"
    return StreamingResponse(
        xml_bytes,
        media_type="application/xml",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# --- SEND REGISTRATION EMAIL ---
async def send_registration_confirmation(person_type: str, person_data: dict, settings: dict):
    if not settings.get("smtp_enabled"):
        return
    club = settings.get("club_name", "Club Deportivo")
    email = person_data.get("email")
    if not email:
        return
    name = f"{person_data.get('name','')} {person_data.get('surname','')}".strip()
    if person_type == "player":
        subject = f"✅ Inscripción confirmada — {club}"
        html = f"""
        <div style="font-family:sans-serif;max-width:520px;margin:0 auto">
          <h2 style="color:#00296B">{club}</h2>
          <h3>Inscripción confirmada</h3>
          <p>Hola <b>{name}</b>,</p>
          <p>Tu inscripción como jugador/a ha sido procesada correctamente.</p>
          <p>El club se pondrá en contacto contigo para completar la documentación.</p>
          <p style="color:#475569;font-size:12px">Si tienes alguna duda, contáctanos en {settings.get('email','')}</p>
        </div>"""
    else:
        member_number = person_data.get("member_number", "")
        subject = f"✅ Alta de socio confirmada — {club}"
        html = f"""
        <div style="font-family:sans-serif;max-width:520px;margin:0 auto">
          <h2 style="color:#00296B">{club}</h2>
          <h3>Alta de socio confirmada</h3>
          <p>Hola <b>{name}</b>,</p>
          <p>Tu alta como socio/a ha sido procesada correctamente.</p>
          {f'<p>Tu número de socio es: <b>{member_number}</b></p>' if member_number else ''}
          <p style="color:#475569;font-size:12px">Si tienes alguna duda, contáctanos en {settings.get('email','')}</p>
        </div>"""
    await send_email(settings, email, subject, html)

@api_router.post("/notifications/send-overdue-reminders")
async def send_overdue_reminders(request: Request):
    """Send email reminders to all people with overdue payments."""
    await get_current_user(request)
    settings = await get_club_settings()
    if not settings.get("smtp_enabled"):
        raise HTTPException(status_code=400, detail="Email SMTP no configurado")

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    overdue = await db.payments.find(
        {"status": "pending", "due_date": {"$lt": today, "$ne": ""}},
        {"_id": 0}
    ).to_list(500)

    sent = 0
    club = settings.get("club_name", "Club Deportivo")
    for payment in overdue:
        person = None
        if payment["person_type"] == "player":
            person = await db.players.find_one({"id": payment["person_id"]}, {"_id": 0})
        else:
            person = await db.members.find_one({"id": payment["person_id"]}, {"_id": 0})
        if not person or not person.get("email"):
            continue
        name = f"{person.get('name','')} {person.get('surname','')}".strip()
        html = f"""
        <div style="font-family:sans-serif;max-width:520px;margin:0 auto">
          <h2 style="color:#00296B">{club}</h2>
          <h3 style="color:#DC2626">Recordatorio de pago pendiente</h3>
          <p>Hola <b>{name}</b>,</p>
          <p>Tienes un pago pendiente:</p>
          <table style="border-collapse:collapse;width:100%">
            <tr><td style="padding:8px;border:1px solid #e2e8f0"><b>Concepto</b></td><td style="padding:8px;border:1px solid #e2e8f0">{payment['concept']}</td></tr>
            <tr><td style="padding:8px;border:1px solid #e2e8f0"><b>Importe</b></td><td style="padding:8px;border:1px solid #e2e8f0">{payment['amount']:.2f} €</td></tr>
            <tr><td style="padding:8px;border:1px solid #e2e8f0"><b>Vencimiento</b></td><td style="padding:8px;border:1px solid #e2e8f0">{payment['due_date']}</td></tr>
          </table>
          <p>Por favor, regulariza tu situación contactando con el club.</p>
          <p style="color:#475569;font-size:12px">{settings.get('email','')} · {settings.get('phone','')}</p>
        </div>"""
        ok = await send_email(settings, person["email"], f"⚠️ Pago pendiente — {club}", html)
        if ok:
            sent += 1

    return {"sent": sent, "total_overdue": len(overdue)}

# --- REPORTS & EXPORT ---

@api_router.get("/reports/summary")
async def reports_summary(request: Request):
    """Dashboard KPIs and aggregated stats for the reports section."""
    await get_current_user(request)

    # Parallel aggregations
    total_players = await db.players.count_documents({})
    active_players = await db.players.count_documents({"status": "active"})
    total_members = await db.members.count_documents({})
    active_members = await db.members.count_documents({"status": "active"})

    # Payments summary
    payments = await db.payments.find({}, {"_id": 0}).to_list(5000)
    total_collected = sum(p["amount"] for p in payments if p.get("status") == "paid")
    total_pending = sum(p["amount"] for p in payments if p.get("status") == "pending")
    overdue_count = 0
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for p in payments:
        if p.get("status") == "pending" and p.get("due_date") and p["due_date"] < today:
            overdue_count += 1

    # Sales by fee
    fees = await db.fees.find({}, {"_id": 0}).to_list(100)
    fee_map = {f["id"]: f for f in fees}
    sales_by_fee = {}
    for p in payments:
        fid = p.get("fee_id", "")
        if fid not in sales_by_fee:
            fee = fee_map.get(fid, {})
            sales_by_fee[fid] = {
                "fee_id": fid,
                "fee_name": fee.get("name", "Sin tarifa"),
                "fee_type": fee.get("fee_type", ""),
                "unit_price": fee.get("amount", 0),
                "count": 0,
                "total_paid": 0.0,
                "total_pending": 0.0,
            }
        sales_by_fee[fid]["count"] += 1
        if p.get("status") == "paid":
            sales_by_fee[fid]["total_paid"] += p["amount"]
        else:
            sales_by_fee[fid]["total_pending"] += p["amount"]

    # Players by team
    teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    team_map = {t["id"]: t["name"] for t in teams}
    players = await db.players.find({}, {"_id": 0}).to_list(5000)
    players_by_team = {}
    for pl in players:
        tid = pl.get("team_id", "")
        tname = team_map.get(tid, "Sin equipo")
        if tname not in players_by_team:
            players_by_team[tname] = {"team": tname, "total": 0, "active": 0, "pending": 0, "inactive": 0}
        players_by_team[tname]["total"] += 1
        st = pl.get("status", "")
        if st == "active":
            players_by_team[tname]["active"] += 1
        elif st == "pending_payment":
            players_by_team[tname]["pending"] += 1
        else:
            players_by_team[tname]["inactive"] += 1

    return {
        "kpis": {
            "total_players": total_players,
            "active_players": active_players,
            "total_members": total_members,
            "active_members": active_members,
            "total_collected": round(total_collected, 2),
            "total_pending": round(total_pending, 2),
            "overdue_count": overdue_count,
        },
        "sales_by_fee": list(sales_by_fee.values()),
        "players_by_team": list(players_by_team.values()),
    }

@api_router.get("/reports/players-detail")
async def reports_players_detail(
    request: Request,
    team_id: Optional[str] = None,
    status: Optional[str] = None,
    fee_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    await get_current_user(request)
    query = {}
    if team_id:
        query["team_id"] = team_id
    if status:
        query["status"] = status
    if date_from or date_to:
        query["created_at"] = {}
        if date_from:
            query["created_at"]["$gte"] = date_from
        if date_to:
            query["created_at"]["$lte"] = date_to + "T23:59:59"
    players = await db.players.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    teams = {t["id"]: t for t in await db.teams.find({}, {"_id": 0}).to_list(100)}
    fees = {f["id"]: f for f in await db.fees.find({}, {"_id": 0}).to_list(100)}
    payments_list = await db.payments.find({"person_type": "player"}, {"_id": 0}).to_list(5000)
    pay_by_player = {}
    for p in payments_list:
        pay_by_player.setdefault(p["person_id"], []).append(p)

    result = []
    for pl in players:
        team = teams.get(pl.get("team_id", ""), {})
        player_payments = pay_by_player.get(pl["id"], [])
        if fee_id:
            if not any(p.get("fee_id") == fee_id for p in player_payments):
                continue
        total_paid = sum(p["amount"] for p in player_payments if p.get("status") == "paid")
        total_pending = sum(p["amount"] for p in player_payments if p.get("status") == "pending")
        result.append({
            "id": pl["id"],
            "name": f"{pl.get('name','')} {pl.get('surname','')}".strip(),
            "dni": pl.get("dni", ""),
            "birthdate": pl.get("birthdate", ""),
            "phone": pl.get("phone", ""),
            "email": pl.get("email", ""),
            "team": team.get("name", "Sin equipo"),
            "category": team.get("category", ""),
            "position": pl.get("position", ""),
            "status": pl.get("status", ""),
            "created_at": pl.get("created_at", "")[:10],
            "total_paid": round(total_paid, 2),
            "total_pending": round(total_pending, 2),
            "payment_status": "pagado" if total_pending == 0 and total_paid > 0 else ("pendiente" if total_pending > 0 else "sin_pago"),
        })
    return result

@api_router.get("/reports/payments-detail")
async def reports_payments_detail(
    request: Request,
    person_type: Optional[str] = None,
    status: Optional[str] = None,
    fee_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    team_id: Optional[str] = None,
):
    await get_current_user(request)
    query = {}
    if person_type:
        query["person_type"] = person_type
    if status:
        query["status"] = status
    if fee_id:
        query["fee_id"] = fee_id
    if date_from or date_to:
        query["created_at"] = {}
        if date_from:
            query["created_at"]["$gte"] = date_from
        if date_to:
            query["created_at"]["$lte"] = date_to + "T23:59:59"
    payments = await db.payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)

    teams = {t["id"]: t["name"] for t in await db.teams.find({}, {"_id": 0}).to_list(100)}
    fees = {f["id"]: f["name"] for f in await db.fees.find({}, {"_id": 0}).to_list(100)}
    players = {p["id"]: p for p in await db.players.find({}, {"_id": 0}).to_list(5000)}
    members = {m["id"]: m for m in await db.members.find({}, {"_id": 0}).to_list(5000)}

    result = []
    for pay in payments:
        person_id = pay.get("person_id", "")
        if pay.get("person_type") == "player":
            person = players.get(person_id, {})
            team_name = teams.get(person.get("team_id", ""), "")
        else:
            person = members.get(person_id, {})
            team_name = ""
        if team_id and person.get("team_id") != team_id:
            continue
        result.append({
            "id": pay["id"],
            "fecha": pay.get("created_at", "")[:10],
            "persona": f"{person.get('name','')} {person.get('surname','')}".strip(),
            "tipo_persona": pay.get("person_type", ""),
            "equipo": team_name,
            "concepto": pay.get("concept", ""),
            "tarifa": fees.get(pay.get("fee_id", ""), ""),
            "importe": pay.get("amount", 0),
            "estado": pay.get("status", ""),
            "metodo": pay.get("notes", "").replace("Pago via ", "").replace("Pago online via ", ""),
            "vencimiento": pay.get("due_date", ""),
            "pagado_el": pay.get("paid_at", "")[:10] if pay.get("paid_at") else "",
        })
    return result

# ─── FACILITIES ──────────────────────────────────────────────────────────────
@api_router.get("/facilities")
async def get_facilities(request: Request):
    await get_current_user(request)
    return await db.facilities.find({}, {"_id": 0}).sort("name", 1).to_list(200)

@api_router.post("/facilities")
async def create_facility(data: FacilityCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.facilities.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/facilities/{fac_id}")
async def update_facility(fac_id: str, data: FacilityUpdate, request: Request):
    await get_current_user(request)
    upd = {k: v for k, v in data.model_dump().items() if v is not None}
    if upd:
        await db.facilities.update_one({"id": fac_id}, {"$set": upd})
    return await db.facilities.find_one({"id": fac_id}, {"_id": 0})

@api_router.delete("/facilities/{fac_id}")
async def delete_facility(fac_id: str, request: Request):
    await get_current_user(request)
    await db.facilities.delete_one({"id": fac_id})
    return {"message": "Instalación eliminada"}


# ─── SALES / VENTAS ──────────────────────────────────────────────────────────
@api_router.get("/sales")
async def get_sales(
    request: Request,
    status: Optional[str] = None,
    person_type: Optional[str] = None,
    payment_method: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    person_id: Optional[str] = None,
):
    await get_current_user(request)
    query: dict = {}
    if status:
        query["status"] = status
    if person_type:
        query["person_type"] = person_type
    if payment_method:
        query["payment_method"] = payment_method
    if person_id:
        query["person_id"] = person_id
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    if date_from or date_to:
        sales = [s for s in sales if (not date_from or s.get("created_at", "")[:10] >= date_from) and
                 (not date_to or s.get("created_at", "")[:10] <= date_to)]
    return sales

@api_router.post("/sales")
async def create_sale(data: SaleCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["paid_at"] = ""
    await db.sales.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/sales/{sale_id}")
async def update_sale(sale_id: str, data: SaleUpdate, request: Request):
    await get_current_user(request)
    upd = {k: v for k, v in data.model_dump().items() if v is not None}
    if upd.get("status") == "paid" and not upd.get("paid_at"):
        upd["paid_at"] = datetime.now(timezone.utc).isoformat()
    if upd:
        await db.sales.update_one({"id": sale_id}, {"$set": upd})
    return await db.sales.find_one({"id": sale_id}, {"_id": 0})

@api_router.delete("/sales/{sale_id}")
async def delete_sale(sale_id: str, request: Request):
    await get_current_user(request)
    await db.sales.delete_one({"id": sale_id})
    return {"message": "Venta eliminada"}

@api_router.post("/sales/bulk")
async def bulk_create_sales(data: BulkSaleRequest, request: Request):
    await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    docs = []
    for pid in data.player_ids:
        docs.append({"id": str(uuid.uuid4()), "person_id": pid, "person_type": "player",
                     "product_id": data.product_id, "fee_id": data.fee_id,
                     "concept": data.concept, "amount": data.amount, "currency": "eur",
                     "payment_method": data.payment_method, "status": "pending",
                     "due_date": data.due_date, "notes": data.notes, "created_at": now, "paid_at": ""})
    for mid in data.member_ids:
        docs.append({"id": str(uuid.uuid4()), "person_id": mid, "person_type": "member",
                     "product_id": data.product_id, "fee_id": data.fee_id,
                     "concept": data.concept, "amount": data.amount, "currency": "eur",
                     "payment_method": data.payment_method, "status": "pending",
                     "due_date": data.due_date, "notes": data.notes, "created_at": now, "paid_at": ""})
    if docs:
        await db.sales.insert_many(docs)
    return {"created": len(docs)}

@api_router.get("/sales/dashboard")
async def sales_dashboard(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    await get_current_user(request)
    sales = await db.sales.find({}, {"_id": 0}).to_list(50000)
    if date_from or date_to:
        sales = [s for s in sales if (not date_from or s.get("created_at", "")[:10] >= date_from) and
                 (not date_to or s.get("created_at", "")[:10] <= date_to)]
    paid = [s for s in sales if s.get("status") == "paid"]
    pending = [s for s in sales if s.get("status") == "pending"]
    cancelled = [s for s in sales if s.get("status") == "cancelled"]
    total_revenue = sum(s.get("amount", 0) for s in paid)
    total_pending = sum(s.get("amount", 0) for s in pending)
    # monthly breakdown
    monthly: dict = {}
    for s in paid:
        month = s.get("created_at", "")[:7]
        if month:
            monthly[month] = monthly.get(month, 0) + s.get("amount", 0)
    # by payment method
    by_method: dict = {}
    for s in paid:
        m = s.get("payment_method", "other")
        by_method[m] = by_method.get(m, 0) + s.get("amount", 0)
    return {
        "total_revenue": round(total_revenue, 2),
        "total_pending": round(total_pending, 2),
        "total_sales": len(sales),
        "paid_count": len(paid),
        "pending_count": len(pending),
        "cancelled_count": len(cancelled),
        "monthly": [{"month": k, "amount": round(v, 2)} for k, v in sorted(monthly.items())],
        "by_method": [{"method": k, "amount": round(v, 2)} for k, v in by_method.items()],
    }

@api_router.get("/sales/export-sepa")
async def export_sales_sepa_list(request: Request, status: Optional[str] = "pending"):
    """Returns sales with payment_method=sepa for SEPA XML generation."""
    await get_current_user(request)
    query: dict = {"payment_method": "sepa"}
    if status:
        query["status"] = status
    sales = await db.sales.find(query, {"_id": 0}).to_list(10000)
    players = {p["id"]: p for p in await db.players.find({}, {"_id": 0}).to_list(5000)}
    members = {m["id"]: m for m in await db.members.find({}, {"_id": 0}).to_list(5000)}
    mandates = {m["person_id"]: m for m in await db.sepa_mandates.find({"status": "active"}, {"_id": 0}).to_list(5000)}
    result = []
    for s in sales:
        pid = s.get("person_id", "")
        person = players.get(pid) or members.get(pid) or {}
        mandate = mandates.get(pid)
        result.append({**s,
            "person_name": f"{person.get('name','')} {person.get('surname','')}".strip(),
            "has_mandate": mandate is not None,
            "mandate_id": mandate["id"] if mandate else "",
            "debtor_iban": mandate["debtor_iban"] if mandate else "",
        })
    return result

@api_router.get("/export/sales")
async def export_sales_excel(
    request: Request,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    payment_method: Optional[str] = None,
):
    await get_current_user(request)
    query: dict = {}
    if status:
        query["status"] = status
    if payment_method:
        query["payment_method"] = payment_method
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(50000)
    if date_from or date_to:
        sales = [s for s in sales if (not date_from or s.get("created_at","")[:10] >= date_from) and
                 (not date_to or s.get("created_at","")[:10] <= date_to)]
    players = {p["id"]: p for p in await db.players.find({}, {"_id": 0}).to_list(5000)}
    members = {m["id"]: m for m in await db.members.find({}, {"_id": 0}).to_list(5000)}
    headers = ["Fecha","Apellidos","Nombre","Tipo","Concepto","Importe €","Estado","Método pago","Vencimiento","Pagado el"]
    rows = []
    for s in sales:
        pid = s.get("person_id","")
        person = players.get(pid) or members.get(pid) or {}
        rows.append([s.get("created_at","")[:10], person.get("surname",""), person.get("name",""),
                     s.get("person_type",""), s.get("concept",""), s.get("amount",0),
                     s.get("status",""), s.get("payment_method",""), s.get("due_date",""),
                     s.get("paid_at","")[:10] if s.get("paid_at") else ""])
    buf = _make_excel("Ventas", headers, rows)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=ventas.xlsx"})


# ─── SCHEDULE EVENTS ─────────────────────────────────────────────────────────
@api_router.get("/schedule/events")
async def get_schedule_events(
    request: Request,
    team_id: Optional[str] = None,
    event_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    await get_current_user(request)
    query: dict = {}
    if team_id:
        query["team_id"] = team_id
    if event_type:
        query["event_type"] = event_type
    events = await db.schedule_events.find(query, {"_id": 0}).sort("date", 1).to_list(5000)
    if date_from:
        events = [e for e in events if e.get("date","") >= date_from]
    if date_to:
        events = [e for e in events if e.get("date","") <= date_to]
    return events

@api_router.post("/schedule/events")
async def create_schedule_event(data: ScheduleEventCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.schedule_events.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.post("/schedule/generate-week")
async def generate_week_from_template(
    request: Request,
    template_id: str,
    week_start: str,  # YYYY-MM-DD (Monday)
):
    """Generate schedule events for a full week from a template."""
    await get_current_user(request)
    tmpl = await db.schedule_templates.find_one({"id": template_id})
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template no encontrado")
    day_map = {"Lunes": 0, "Martes": 1, "Miércoles": 2, "Jueves": 3, "Viernes": 4, "Sábado": 5, "Domingo": 6}
    dow = day_map.get(tmpl.get("day_of_week", "Lunes"), 0)
    from datetime import timedelta
    base = datetime.strptime(week_start, "%Y-%m-%d")
    event_date = base + timedelta(days=dow)
    repeat = int(tmpl.get("repeat_weeks", 1))
    created = []
    for w in range(repeat):
        d = event_date + timedelta(weeks=w)
        doc = {
            "id": str(uuid.uuid4()),
            "title": tmpl.get("name", "Entrenamiento"),
            "team_id": tmpl.get("team_id", ""),
            "facility_id": tmpl.get("facility_id", ""),
            "event_type": tmpl.get("event_type", "training"),
            "date": d.strftime("%Y-%m-%d"),
            "start_time": tmpl.get("start_time", ""),
            "end_time": tmpl.get("end_time", ""),
            "notes": tmpl.get("notes", ""),
            "category": "", "opponent": "", "home_away": "home", "player_ids": [],
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.schedule_events.insert_one(doc)
        created.append({k: v for k, v in doc.items() if k != "_id"})
    return {"created": len(created), "events": created}

@api_router.get("/schedule/templates")
async def get_schedule_templates(request: Request):
    await get_current_user(request)
    return await db.schedule_templates.find({}, {"_id": 0}).sort("name", 1).to_list(200)

@api_router.post("/schedule/templates")
async def create_schedule_template(data: ScheduleTemplateCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.schedule_templates.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.post("/schedule/templates/{tmpl_id}/generate-week")
async def generate_week_from_template_v2(tmpl_id: str, body: dict, request: Request):
    """Generate schedule events for a full week from a multi-event template."""
    await get_current_user(request)
    tmpl = await db.schedule_templates.find_one({"id": tmpl_id}, {"_id": 0})
    if not tmpl:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    week_start_str = body.get("week_start", "")
    if not week_start_str:
        raise HTTPException(status_code=400, detail="week_start requerido")
    base = datetime.strptime(week_start_str, "%Y-%m-%d")
    template_events = tmpl.get("events", [])
    # Support legacy single-event templates
    if not template_events and tmpl.get("day_of_week"):
        day_map = {"Lunes": 0, "Martes": 1, "Miércoles": 2, "Jueves": 3, "Viernes": 4, "Sábado": 5, "Domingo": 6}
        template_events = [{
            "type": tmpl.get("event_type", "entrenamiento"),
            "title": tmpl.get("name", ""),
            "team_id": tmpl.get("team_id", ""),
            "facility_id": tmpl.get("facility_id", ""),
            "day_of_week": day_map.get(tmpl.get("day_of_week", "Lunes"), 0),
            "time": tmpl.get("start_time", "09:00"),
            "duration_min": 90,
        }]
    from datetime import timedelta
    created = 0
    for ev in template_events:
        dow = ev.get("day_of_week", 0)
        event_date = base + timedelta(days=int(dow))
        doc = {
            "id": str(uuid.uuid4()),
            "type": ev.get("type", "entrenamiento"),
            "title": ev.get("title") or ev.get("type", "entrenamiento"),
            "team_id": ev.get("team_id", ""),
            "facility_id": ev.get("facility_id", ""),
            "date": event_date.strftime("%Y-%m-%d"),
            "time": ev.get("time", "09:00"),
            "duration_min": ev.get("duration_min", 90),
            "notes": "",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.schedule_events.insert_one(doc)
        created += 1
    return {"created": created}

@api_router.delete("/schedule/events/{event_id}")
async def delete_schedule_event(event_id: str, request: Request):
    await get_current_user(request)
    await db.schedule_events.delete_one({"id": event_id})
    return {"message": "Evento eliminado"}

@api_router.delete("/schedule/templates/{tmpl_id}")
async def delete_schedule_template(tmpl_id: str, request: Request):
    await get_current_user(request)
    await db.schedule_templates.delete_one({"id": tmpl_id})
    return {"message": "Plantilla eliminada"}


# ─── ENHANCED DASHBOARD ───────────────────────────────────────────────────────
@api_router.get("/dashboard/stats")
async def dashboard_stats(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    await get_current_user(request)
    now_str = datetime.now(timezone.utc).isoformat()
    if not date_from:
        date_from = datetime.now(timezone.utc).replace(day=1).strftime("%Y-%m-%d")
    if not date_to:
        date_to = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    all_sales = await db.sales.find({}, {"_id": 0}).to_list(50000)
    period_sales = [s for s in all_sales if date_from <= s.get("created_at","")[:10] <= date_to]
    paid_period = [s for s in period_sales if s.get("status") == "paid"]
    pending_all = [s for s in all_sales if s.get("status") == "pending"]
    overdue_all = [s for s in pending_all if s.get("due_date","") and s.get("due_date","") < now_str[:10]]

    players = await db.players.find({}, {"_id": 0}).to_list(5000)
    members = await db.members.find({}, {"_id": 0}).to_list(5000)
    teams = await db.teams.find({}, {"_id": 0}).to_list(200)

    # Monthly revenue last 6 months
    monthly: dict = {}
    for s in [x for x in all_sales if x.get("status") == "paid"]:
        m = s.get("created_at","")[:7]
        if m:
            monthly[m] = monthly.get(m, 0) + s.get("amount", 0)
    sorted_months = sorted(monthly.items())[-6:]

    # Recent sales
    recent = sorted(period_sales, key=lambda x: x.get("created_at",""), reverse=True)[:10]
    players_map = {p["id"]: f"{p.get('name','')} {p.get('surname','')}".strip() for p in players}
    members_map = {m["id"]: f"{m.get('name','')} {m.get('surname','')}".strip() for m in members}
    for s in recent:
        pid = s.get("person_id","")
        s["person_name"] = players_map.get(pid) or members_map.get(pid) or "—"

    return {
        "period": {"from": date_from, "to": date_to},
        "revenue": round(sum(s.get("amount",0) for s in paid_period), 2),
        "revenue_count": len(paid_period),
        "pending_total": round(sum(s.get("amount",0) for s in pending_all), 2),
        "pending_count": len(pending_all),
        "overdue_total": round(sum(s.get("amount",0) for s in overdue_all), 2),
        "overdue_count": len(overdue_all),
        "players_total": len(players),
        "players_active": len([p for p in players if p.get("status") == "active"]),
        "members_total": len(members),
        "members_active": len([m for m in members if m.get("status") == "active"]),
        "teams_total": len(teams),
        "monthly_revenue": [{"month": m, "amount": round(v, 2)} for m, v in sorted_months],
        "recent_sales": recent,
    }


# ─── PRODUCTS ──────────────────────────────────────────────────────────────────
@api_router.get("/products")
async def get_products(request: Request, active_only: bool = False):
    await get_current_user(request)
    query = {"active": True} if active_only else {}
    items = await db.products.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return items

@api_router.post("/products")
async def create_product(data: ProductCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.products.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/products/{product_id}")
async def update_product(product_id: str, data: ProductUpdate, request: Request):
    await get_current_user(request)
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.products.update_one({"id": product_id}, {"$set": update_data})
    return await db.products.find_one({"id": product_id}, {"_id": 0})

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, request: Request):
    await get_current_user(request)
    await db.products.delete_one({"id": product_id})
    return {"message": "Producto eliminado"}


# ─── ADMIN USERS ───────────────────────────────────────────────────────────────
@api_router.get("/admin/users")
async def get_admin_users(request: Request):
    user = await get_current_user(request)
    if user.get("role") not in ("admin", "director"):
        raise HTTPException(status_code=403, detail="Sin permisos")
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(200)
    for u in users:
        if "_id" in u:
            u["_id"] = str(u["_id"])
    return users

@api_router.post("/admin/users")
async def create_admin_user(data: AdminUserCreate, request: Request):
    user = await get_current_user(request)
    if user.get("role") not in ("admin", "director"):
        raise HTTPException(status_code=403, detail="Sin permisos")
    existing = await db.users.find_one({"email": data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email ya en uso")
    doc = {
        "id": str(uuid.uuid4()),
        "email": data.email.lower(),
        "name": data.name,
        "password_hash": hash_password(data.password),
        "role": data.role,
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(doc)
    doc.pop("password_hash", None)
    doc.pop("_id", None)
    return doc

@api_router.put("/admin/users/{user_id}")
async def update_admin_user(user_id: str, data: AdminUserUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") not in ("admin", "director"):
        raise HTTPException(status_code=403, detail="Sin permisos")
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if "password" in update_data:
        update_data["password_hash"] = hash_password(update_data.pop("password"))
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return updated

@api_router.delete("/admin/users/{user_id}")
async def delete_admin_user(user_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") not in ("admin", "director"):
        raise HTTPException(status_code=403, detail="Sin permisos")
    if user.get("id") == user_id:
        raise HTTPException(status_code=400, detail="No puedes eliminar tu propio usuario")
    await db.users.delete_one({"id": user_id})
    return {"message": "Usuario eliminado"}


# ─── EXCEL EXPORT HELPERS ─────────────────────────────────────────────────────
def _make_excel(title: str, headers: list, rows: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_fill = PatternFill("solid", fgColor="00296B")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── EXPORT ENDPOINTS ─────────────────────────────────────────────────────────
@api_router.get("/export/players")
async def export_players(
    request: Request,
    team_id: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    birth_year: Optional[int] = None,
    has_siblings: Optional[bool] = None,
    is_minor: Optional[bool] = None,
):
    await get_current_user(request)
    query = {}
    if team_id:
        query["team_id"] = team_id
    if status:
        query["status"] = status
    players = await db.players.find(query, {"_id": 0}).sort("surname", 1).to_list(5000)
    teams = {t["id"]: t["name"] for t in await db.teams.find({}, {"_id": 0}).to_list(200)}

    # Detect siblings: group by family_id
    family_counts: dict = {}
    for p in players:
        fid = p.get("family_id", "")
        if fid:
            family_counts[fid] = family_counts.get(fid, 0) + 1

    now = datetime.now()
    rows = []
    for p in players:
        if birth_year:
            bdate = p.get("birthdate", "")
            p_year = int(bdate[:4]) if bdate and len(bdate) >= 4 else 0
            if p_year != birth_year:
                continue
        if category:
            if teams.get(p.get("team_id", ""), "") != category:
                continue
        fid = p.get("family_id", "")
        siblings = (family_counts.get(fid, 0) - 1) if fid else 0
        if has_siblings is not None:
            if has_siblings and siblings < 1:
                continue
            if not has_siblings and siblings > 0:
                continue
        bdate = p.get("birthdate", "")
        age = None
        if bdate:
            try:
                bd = datetime.strptime(bdate[:10], "%Y-%m-%d")
                age = (now - bd).days // 365
            except Exception:
                pass
        if is_minor is not None:
            if is_minor and (age is None or age >= 18):
                continue
            if not is_minor and (age is None or age < 18):
                continue
        rows.append([
            p.get("surname", ""), p.get("name", ""),
            p.get("birthdate", ""), age,
            p.get("dni", ""), p.get("phone", ""), p.get("email", ""),
            teams.get(p.get("team_id", ""), ""), p.get("status", ""),
            p.get("city", ""), p.get("address", ""), p.get("postal_code", ""),
            p.get("medical_notes", ""), p.get("blood_type", ""),
            siblings, p.get("family_id", ""),
            p.get("season", ""), p.get("notes", ""),
        ])
    headers = [
        "Apellidos", "Nombre", "F.Nacimiento", "Edad",
        "NIF/NIE/Pasaporte", "Teléfono", "Email",
        "Equipo", "Estado",
        "Ciudad", "Dirección", "CP",
        "Notas médicas", "Grupo sanguíneo",
        "Hermanos en club", "ID Familiar",
        "Temporada", "Observaciones",
    ]
    buf = _make_excel("Deportistas", headers, rows)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=deportistas.xlsx"},
    )


@api_router.get("/export/members")
async def export_members(
    request: Request,
    member_type: Optional[str] = None,
    status: Optional[str] = None,
):
    await get_current_user(request)
    query = {}
    if member_type:
        query["member_type"] = member_type
    if status:
        query["status"] = status
    members = await db.members.find(query, {"_id": 0}).sort("surname", 1).to_list(5000)
    headers = [
        "Nº Socio", "Apellidos", "Nombre", "Tipo",
        "NIF/NIE", "F.Nacimiento", "Teléfono", "Email",
        "Dirección", "CP", "Ciudad",
        "IBAN", "Estado", "Temporada", "Observaciones",
    ]
    rows = [[
        m.get("member_number", ""),
        m.get("surname", ""), m.get("name", ""),
        m.get("member_type", ""), m.get("dni", ""), m.get("birthdate", ""),
        m.get("phone", ""), m.get("email", ""),
        m.get("address", ""), m.get("postal_code", ""), m.get("city", ""),
        m.get("bank_iban", ""), m.get("status", ""), m.get("season", ""),
        m.get("notes", ""),
    ] for m in members]
    buf = _make_excel("Socios", headers, rows)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=socios.xlsx"},
    )


@api_router.get("/export/payments")
async def export_payments(
    request: Request,
    status: Optional[str] = None,
    person_type: Optional[str] = None,
    team_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    if person_type:
        query["person_type"] = person_type
    payments = await db.payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    players = {p["id"]: p for p in await db.players.find({}, {"_id": 0}).to_list(5000)}
    members = {m["id"]: m for m in await db.members.find({}, {"_id": 0}).to_list(5000)}
    teams = {t["id"]: t["name"] for t in await db.teams.find({}, {"_id": 0}).to_list(200)}
    fees = {f["id"]: f["name"] for f in await db.fees.find({}, {"_id": 0}).to_list(200)}
    headers = [
        "Fecha", "Apellidos", "Nombre", "Tipo Persona", "Equipo",
        "Concepto", "Tarifa", "Importe (€)", "Estado",
        "Vencimiento", "Pagado el", "Método",
    ]
    rows = []
    for pay in payments:
        fecha = pay.get("created_at", "")[:10]
        if date_from and fecha < date_from:
            continue
        if date_to and fecha > date_to:
            continue
        pid = pay.get("person_id", "")
        if pay.get("person_type") == "player":
            person = players.get(pid, {})
            team = teams.get(person.get("team_id", ""), "")
        else:
            person = members.get(pid, {})
            team = ""
        if team_id and person.get("team_id") != team_id:
            continue
        rows.append([
            fecha,
            person.get("surname", ""), person.get("name", ""),
            pay.get("person_type", ""), team,
            pay.get("concept", ""), fees.get(pay.get("fee_id", ""), ""),
            pay.get("amount", 0), pay.get("status", ""),
            pay.get("due_date", ""),
            pay.get("paid_at", "")[:10] if pay.get("paid_at") else "",
            pay.get("notes", ""),
        ])
    buf = _make_excel("Pagos", headers, rows)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pagos.xlsx"},
    )

@api_router.get("/export/guardians")
async def export_guardians(request: Request):
    await get_current_user(request)
    guardians = await db.guardians.find({}, {"_id": 0}).sort("surname", 1).to_list(5000)
    players = {p["id"]: f"{p.get('name','')} {p.get('surname','')}".strip()
               for p in await db.players.find({}, {"_id": 0}).to_list(5000)}
    headers = [
        "Apellidos", "Nombre", "Relación", "NIF/NIE",
        "Teléfono", "Email", "Dirección", "Ciudad",
        "IBAN", "Jugadores a cargo", "Observaciones",
    ]
    rows = [[
        g.get("surname", ""), g.get("name", ""), g.get("relationship", ""),
        g.get("dni", ""), g.get("phone", ""), g.get("email", ""),
        g.get("address", ""), g.get("city", ""), g.get("bank_iban", ""),
        " | ".join([players.get(pid, pid) for pid in g.get("player_ids", [])]),
        g.get("notes", ""),
    ] for g in guardians]
    buf = _make_excel("Tutores", headers, rows)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tutores.xlsx"},
    )


# ─── COMMUNICATIONS LISTS ─────────────────────────────────────────────────────
@api_router.get("/communications/lists")
async def get_comm_lists(request: Request):
    await get_current_user(request)
    lists = await db.comm_lists.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    return lists

@api_router.post("/communications/lists")
async def create_comm_list(data: CommListCreate, request: Request):
    await get_current_user(request)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.comm_lists.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/communications/lists/{list_id}")
async def update_comm_list(list_id: str, data: CommListUpdate, request: Request):
    await get_current_user(request)
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.comm_lists.update_one({"id": list_id}, {"$set": update_data})
    return await db.comm_lists.find_one({"id": list_id}, {"_id": 0})

@api_router.delete("/communications/lists/{list_id}")
async def delete_comm_list(list_id: str, request: Request):
    await get_current_user(request)
    await db.comm_lists.delete_one({"id": list_id})
    return {"message": "Lista eliminada"}


async def _resolve_list_emails(list_id: str) -> list:
    """Resolve a comm list to a deduplicated list of email addresses."""
    lst = await db.comm_lists.find_one({"id": list_id})
    if not lst:
        return []
    ft = lst.get("filter_type", "manual")
    emails = set()
    if ft == "manual":
        rtype = lst.get("recipient_type", "player")
        for rid in lst.get("recipient_ids", []):
            coll = db.players if rtype == "player" else db.members if rtype == "member" else db.guardians
            doc = await coll.find_one({"id": rid})
            if doc and doc.get("email"):
                emails.add(doc["email"])
    elif ft == "all_players":
        async for p in db.players.find({"status": "active"}, {"email": 1}):
            if p.get("email"):
                emails.add(p["email"])
    elif ft == "all_members":
        async for m in db.members.find({"status": "active"}, {"email": 1}):
            if m.get("email"):
                emails.add(m["email"])
    elif ft == "all_guardians":
        async for g in db.guardians.find({}, {"email": 1}):
            if g.get("email"):
                emails.add(g["email"])
    elif ft == "team":
        team_id_val = lst.get("filter_value", "")
        async for p in db.players.find({"team_id": team_id_val, "status": "active"}, {"email": 1}):
            if p.get("email"):
                emails.add(p["email"])
    elif ft == "category":
        cat = lst.get("filter_value", "")
        teams = await db.teams.find({"category": cat}, {"id": 1}).to_list(50)
        tids = [t["id"] for t in teams]
        async for p in db.players.find({"team_id": {"$in": tids}, "status": "active"}, {"email": 1}):
            if p.get("email"):
                emails.add(p["email"])
    return list(emails)


@api_router.post("/communications/send")
async def send_communication(data: CommSendRequest, request: Request):
    user = await get_current_user(request)
    settings = await db.settings.find_one({}, {"_id": 0}) or {}
    if not settings.get("smtp_enabled"):
        raise HTTPException(status_code=400, detail="SMTP no configurado. Actívalo en Ajustes → Email SMTP.")

    # Resolve recipients
    all_emails: list = list(data.recipient_emails or [])
    if data.list_id:
        all_emails += await _resolve_list_emails(data.list_id)
    all_emails = list(set(e.strip() for e in all_emails if e.strip()))
    if not all_emails:
        raise HTTPException(status_code=400, detail="No hay destinatarios con email válido")

    # Send via SMTP
    sent = 0
    errors = []
    try:
        smtp_host = settings["smtp_host"]
        smtp_port = int(settings.get("smtp_port", 587))
        smtp_user = settings["smtp_user"]
        smtp_pass = settings["smtp_password"]
        from_name = settings.get("smtp_from_name", settings["smtp_user"])
        use_tls = settings.get("smtp_use_tls", True)

        conn = smtplib.SMTP(smtp_host, smtp_port, timeout=15)
        if use_tls:
            conn.starttls()
        conn.login(smtp_user, smtp_pass)

        for email_addr in all_emails:
            try:
                msg = MIMEMultipart("alternative")
                msg["Subject"] = data.subject
                msg["From"] = f"{from_name} <{smtp_user}>"
                msg["To"] = email_addr
                msg.attach(MIMEText(data.body_html, "html", "utf-8"))
                conn.sendmail(smtp_user, [email_addr], msg.as_string())
                sent += 1
            except Exception as e:
                errors.append({"email": email_addr, "error": str(e)})
        conn.quit()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error SMTP: {str(e)}")

    # Save to history
    history_doc = {
        "id": str(uuid.uuid4()),
        "subject": data.subject,
        "body_html": data.body_html,
        "list_id": data.list_id or "",
        "recipient_count": len(all_emails),
        "sent_count": sent,
        "errors": errors,
        "sent_by": user.get("email", ""),
        "sent_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.comm_history.insert_one(history_doc)
    return {"sent": sent, "total": len(all_emails), "errors": errors}


@api_router.post("/communications/send-with-attachments")
async def send_communication_with_attachments(
    request: Request,
    subject: str = Form(...),
    body_html: str = Form(...),
    list_id: str = Form(""),
    recipient_emails: list[str] = Form(default=[]),
    attachments: list[UploadFile] = File(default=[]),
):
    user = await get_current_user(request)
    settings = await db.settings.find_one({}, {"_id": 0}) or {}
    if not settings.get("smtp_enabled"):
        raise HTTPException(status_code=400, detail="SMTP no configurado. Actívalo en Ajustes → Email SMTP.")

    all_emails: list = list(recipient_emails or [])
    if list_id:
        all_emails += await _resolve_list_emails(list_id)
    all_emails = list(set(e.strip() for e in all_emails if e.strip()))
    if not all_emails:
        raise HTTPException(status_code=400, detail="No hay destinatarios con email válido")

    # Pre-read all attachment data
    attachment_data = []
    for att in attachments:
        content = await att.read()
        attachment_data.append({"filename": att.filename, "content": content, "content_type": att.content_type or "application/octet-stream"})

    smtp_host = settings["smtp_host"]
    smtp_port = int(settings.get("smtp_port", 587))
    smtp_user = settings["smtp_user"]
    smtp_pass = settings["smtp_password"]
    from_name = settings.get("smtp_from_name", smtp_user)
    use_tls = settings.get("smtp_use_tls", True)

    sent = 0
    errors = []
    try:
        conn = smtplib.SMTP(smtp_host, smtp_port, timeout=15)
        if use_tls:
            conn.starttls()
        conn.login(smtp_user, smtp_pass)

        for email_addr in all_emails:
            try:
                msg = MIMEMultipart("mixed")
                msg["Subject"] = subject
                msg["From"] = f"{from_name} <{smtp_user}>"
                msg["To"] = email_addr
                msg.attach(MIMEText(body_html, "html", "utf-8"))
                for att in attachment_data:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(att["content"])
                    encoders.encode_base64(part)
                    part.add_header("Content-Disposition", "attachment", filename=att["filename"])
                    msg.attach(part)
                conn.sendmail(smtp_user, [email_addr], msg.as_string())
                sent += 1
            except Exception as e:
                errors.append({"email": email_addr, "error": str(e)})
        conn.quit()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error SMTP: {str(e)}")

    history_doc = {
        "id": str(uuid.uuid4()),
        "subject": subject,
        "body_html": body_html,
        "list_id": list_id or "",
        "attachment_names": [a["filename"] for a in attachment_data],
        "recipient_count": len(all_emails),
        "sent_count": sent,
        "errors": errors,
        "sent_by": user.get("email", ""),
        "sent_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.comm_history.insert_one(history_doc)
    return {"sent": sent, "total": len(all_emails), "errors": errors}


@api_router.get("/communications/history")
async def get_comm_history(request: Request, limit: int = 50):
    await get_current_user(request)
    history = await db.comm_history.find({}, {"_id": 0}).sort("sent_at", -1).to_list(limit)
    return history

@api_router.get("/communications/preview-list/{list_id}")
async def preview_comm_list(list_id: str, request: Request):
    """Returns how many emails a list would resolve to."""
    await get_current_user(request)
    emails = await _resolve_list_emails(list_id)
    return {"count": len(emails), "emails": emails[:20]}  # preview max 20


# Health check
@api_router.get("/health")
async def health():
    return {"status": "ok"}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.players.create_index("team_id")
    await db.players.create_index("family_id")
    await db.guardians.create_index("player_ids")
    await db.guardians.create_index("family_id")
    await db.members.create_index("member_number", unique=True, sparse=True)
    await db.training_schedules.create_index("team_id")
    await db.fees.create_index("fee_type")
    await db.payments.create_index("person_id")
    await db.payments.create_index("status")
    await db.payments.create_index("stripe_session_id")
    await db.sepa_mandates.create_index("person_id")
    await db.sepa_mandates.create_index("status")
    await db.products.create_index("category")
    await db.comm_lists.create_index("filter_type")
    await db.comm_history.create_index("sent_at")
    await db.facilities.create_index("name")
    await db.sales.create_index("person_id")
    await db.sales.create_index("status")
    await db.sales.create_index("created_at")
    await db.schedule_events.create_index("date")
    await db.schedule_events.create_index("team_id")
    await db.schedule_templates.create_index("team_id")
    await seed_admin()
    await seed_demo_data()

async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@racingsangabriel.es")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Racing2025!")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        hashed = hash_password(admin_password)
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hashed,
            "name": "Administrador",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin seeded: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
        logger.info("Admin password updated")

async def seed_demo_data():
    # Seed teams if empty
    count = await db.teams.count_documents({})
    if count == 0:
        teams = [
            {"id": str(uuid.uuid4()), "name": "Alevin A", "category": "Alevin", "coach": "", "image_url": "", "description": "Equipo alevin categoria A", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Alevin B", "category": "Alevin", "coach": "", "image_url": "", "description": "Equipo alevin categoria B", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Benjamin A", "category": "Benjamin", "coach": "", "image_url": "", "description": "Equipo benjamin categoria A", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Benjamin B", "category": "Benjamin", "coach": "", "image_url": "", "description": "Equipo benjamin categoria B", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Cadete", "category": "Cadete", "coach": "", "image_url": "", "description": "Equipo cadete", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Escuela Iniciacion", "category": "Escuela", "coach": "", "image_url": "", "description": "Escuela de iniciacion al futbol", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Futbol Femenino", "category": "Femenino", "coach": "", "image_url": "", "description": "4 equipos de futbol femenino", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Futbol Sala Senior", "category": "Futbol Sala", "coach": "", "image_url": "", "description": "Equipo de futbol sala senior", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Infantil A", "category": "Infantil", "coach": "", "image_url": "", "description": "Equipo infantil categoria A", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Infantil B", "category": "Infantil", "coach": "", "image_url": "", "description": "Equipo infantil categoria B", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Juvenil A", "category": "Juvenil", "coach": "", "image_url": "https://customer-assets.emergentagent.com/job_sg-racing-portal/artifacts/twkjfnq4_image.png", "description": "Equipo juvenil categoria A", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Juvenil B", "category": "Juvenil", "coach": "", "image_url": "", "description": "Equipo juvenil categoria B", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Prebenjamin", "category": "Prebenjamin", "coach": "", "image_url": "", "description": "Equipo prebenjamin", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Senior", "category": "Senior", "coach": "", "image_url": "", "description": "Equipo senior", "created_at": datetime.now(timezone.utc).isoformat()},
        ]
        await db.teams.insert_many(teams)

    # Seed matches if empty
    count = await db.matches.count_documents({})
    if count == 0:
        matches = [
            {"id": str(uuid.uuid4()), "home_team": "Racing San Gabriel", "away_team": "CD Alicante", "date": "2025-02-15", "time": "10:30", "location": "Campo Municipal San Gabriel", "category": "Juvenil", "result": "2-1", "status": "played", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "home_team": "Racing San Gabriel", "away_team": "CF Intercity B", "date": "2025-02-22", "time": "12:00", "location": "Campo Municipal San Gabriel", "category": "Juvenil", "result": "", "status": "upcoming", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "home_team": "Hércules CF C", "away_team": "Racing San Gabriel", "date": "2025-03-01", "time": "11:00", "location": "Estadio Rico Pérez", "category": "Cadete", "result": "", "status": "upcoming", "created_at": datetime.now(timezone.utc).isoformat()},
        ]
        await db.matches.insert_many(matches)

    # Seed news if empty
    count = await db.news.count_documents({})
    if count == 0:
        news = [
            {"id": str(uuid.uuid4()), "title": "Inscripciones abiertas temporada 2025/2026", "content": "Ya puedes inscribirte en cualquiera de nuestras categorias: desde Prebenjamin hasta Senior, pasando por Futbol Femenino y Futbol Sala. Contactanos en el 617 50 27 80 o en racingsangabrieladc@hotmail.com", "image_url": "https://customer-assets.emergentagent.com/job_sg-racing-portal/artifacts/twkjfnq4_image.png", "source": "web", "category": "general", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "title": "Futbol femenino en auge", "content": "Nuestro club cuenta ya con 4 equipos de futbol femenino. Seguimos creciendo y apostando por el deporte femenino en Alicante. Si quieres unirte, contactanos.", "image_url": "https://images.unsplash.com/photo-1652190416554-c46af8a0ff50?crop=entropy&cs=srgb&fm=jpg&ixid=M3w3NDQ2NDN8MHwxfHNlYXJjaHw0fHxzb2NjZXIlMjBmb290YmFsbCUyMGZpZWxkJTIwdHJhaW5pbmd8ZW58MHx8fHwxNzc2MjY4MjIwfDA&ixlib=rb-4.1.0&q=85", "source": "web", "category": "general", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "title": "Nuevas actividades en la Sala Multiactividad", "content": "Ademas de futbol, en nuestra sede ofrecemos clases de Zumba, Fitness y Pilates en la Sala Multiactividad. Consulta horarios en la sede.", "image_url": "", "source": "web", "category": "eventos", "created_at": datetime.now(timezone.utc).isoformat()},
        ]
        await db.news.insert_many(news)

    # Seed gallery if empty
    count = await db.gallery.count_documents({})
    if count == 0:
        gallery = [
            {"id": str(uuid.uuid4()), "title": "Equipo Juvenil A 2024/2025", "image_url": "https://customer-assets.emergentagent.com/job_sg-racing-portal/artifacts/twkjfnq4_image.png", "description": "Foto oficial del equipo juvenil", "category": "equipos", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "title": "Partido en casa", "image_url": "https://images.pexels.com/photos/36213470/pexels-photo-36213470.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940", "description": "Acción durante partido de liga", "category": "partidos", "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "title": "Estadio", "image_url": "https://static.prod-images.emergentagent.com/jobs/aa4aac70-2da7-49b9-b970-59a86d8b85ed/images/2e97b7f318b408c34d0d83e3483ffb1fb1d9ea389c42fb72c00f0fdd9219dad4.png", "description": "Vista del estadio", "category": "instalaciones", "created_at": datetime.now(timezone.utc).isoformat()},
        ]
        await db.gallery.insert_many(gallery)

    # Seed sample social posts (to show how it looks before n8n is connected)
    count = await db.social_posts.count_documents({})
    if count == 0:
        social_posts = [
            {"id": str(uuid.uuid4()), "source": "instagram", "content": "Gran entrenamiento de nuestro equipo Juvenil A esta tarde. Preparandose para el proximo partido de liga. Vamos Racing!", "image_url": "https://customer-assets.emergentagent.com/job_sg-racing-portal/artifacts/twkjfnq4_image.png", "post_url": "https://www.instagram.com/racingsangabrieladc/", "author": "@racingsangabrieladc", "posted_at": datetime.now(timezone.utc).isoformat(), "received_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "source": "instagram", "content": "Inscripciones abiertas para la temporada 2025/2026. Todas las categorias disponibles. Contactanos!", "image_url": "https://images.unsplash.com/photo-1652190416554-c46af8a0ff50?crop=entropy&cs=srgb&fm=jpg&ixid=M3w3NDQ2NDN8MHwxfHNlYXJjaHw0fHxzb2NjZXIlMjBmb290YmFsbCUyMGZpZWxkJTIwdHJhaW5pbmd8ZW58MHx8fHwxNzc2MjY4MjIwfDA&ixlib=rb-4.1.0&q=85", "post_url": "https://www.instagram.com/racingsangabrieladc/", "author": "@racingsangabrieladc", "posted_at": datetime.now(timezone.utc).isoformat(), "received_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "source": "instagram", "content": "Nuestro futbol femenino sigue creciendo. Ya contamos con 4 equipos!", "image_url": "", "post_url": "https://www.instagram.com/racingsangabrieladc/", "author": "@racingsangabrieladc", "posted_at": datetime.now(timezone.utc).isoformat(), "received_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "source": "instagram", "content": "Clase de Zumba esta tarde en la Sala Multiactividad. Te apuntas?", "image_url": "", "post_url": "https://www.instagram.com/racingsangabrieladc/", "author": "@racingsangabrieladc", "posted_at": datetime.now(timezone.utc).isoformat(), "received_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "source": "facebook", "content": "Victoria del Alevin A este fin de semana. Gran trabajo de todo el equipo y del cuerpo tecnico. Seguimos sumando!", "image_url": "https://customer-assets.emergentagent.com/job_sg-racing-portal/artifacts/twkjfnq4_image.png", "post_url": "https://www.facebook.com/RacingSanGabrielADC/", "author": "Racing San Gabriel ADC", "posted_at": datetime.now(timezone.utc).isoformat(), "received_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "source": "facebook", "content": "Horario de entrenamientos actualizado para esta semana. Consulta en la sede o llamanos al 617 50 27 80.", "image_url": "", "post_url": "https://www.facebook.com/RacingSanGabrielADC/", "author": "Racing San Gabriel ADC", "posted_at": datetime.now(timezone.utc).isoformat(), "received_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "source": "facebook", "content": "Nuevas equipaciones disponibles para todos los equipos. Pasate por la sede!", "image_url": "https://images.unsplash.com/photo-1652190416150-d501a60291b3?crop=entropy&cs=srgb&fm=jpg&ixid=M3w3NDQ2NDN8MHwxfHNlYXJjaHwyfHxzb2NjZXIlMjBmb290YmFsbCUyMGZpZWxkJTIwdHJhaW5pbmd8ZW58MHx8fHwxNzc2MjY4MjIwfDA&ixlib=rb-4.1.0&q=85", "post_url": "https://www.facebook.com/RacingSanGabrielADC/", "author": "Racing San Gabriel ADC", "posted_at": datetime.now(timezone.utc).isoformat(), "received_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "source": "facebook", "content": "Torneo de verano para categorias base. Pronto mas informacion!", "image_url": "", "post_url": "https://www.facebook.com/RacingSanGabrielADC/", "author": "Racing San Gabriel ADC", "posted_at": datetime.now(timezone.utc).isoformat(), "received_at": datetime.now(timezone.utc).isoformat()},
        ]
        await db.social_posts.insert_many(social_posts)

    # Seed default fees (cuotas de socios RSG)
    count = await db.fees.count_documents({})
    if count == 0:
        now_iso = datetime.now(timezone.utc).isoformat()
        fees = [
            {
                "id": str(uuid.uuid4()), "name": "Cuota Socio Sede Club RSG",
                "amount": 32.50, "fee_type": "socio", "period": "trimestral",
                "description": "Cuota trimestral para socios de la sede del club",
                "active": True, "created_at": now_iso,
            },
            {
                "id": str(uuid.uuid4()), "name": "Cuota Socio Reducida RSG",
                "amount": 30.00, "fee_type": "socio", "period": "trimestral",
                "description": "Cuota trimestral reducida (familias numerosas, desempleados...)",
                "active": True, "created_at": now_iso,
            },
            {
                "id": str(uuid.uuid4()), "name": "Cuota Socio Menor RSG",
                "amount": 10.00, "fee_type": "socio", "period": "trimestral",
                "description": "Cuota trimestral para socios menores de edad",
                "active": True, "created_at": now_iso,
            },
        ]
        await db.fees.insert_many(fees)
        logger.info("Default fees seeded (32.50€, 30€, 10€ trimestral)")

    # Seed default products (equipaciones, material)
    count = await db.products.count_documents({})
    if count == 0:
        now_iso = datetime.now(timezone.utc).isoformat()
        products = [
            {
                "id": str(uuid.uuid4()), "name": "Kit Equipación Completa", "category": "equipacion",
                "price": 45.00, "stock": None, "active": True,
                "description": "Camiseta + pantalón + medias oficiales del club",
                "is_recurring": False, "created_at": now_iso,
            },
            {
                "id": str(uuid.uuid4()), "name": "Camiseta Oficial Racing San Gabriel", "category": "equipacion",
                "price": 20.00, "stock": None, "active": True,
                "description": "Camiseta oficial temporada 26/27",
                "is_recurring": False, "created_at": now_iso,
            },
            {
                "id": str(uuid.uuid4()), "name": "Cuota Mensual Actividades", "category": "cuota",
                "price": 25.00, "stock": None, "active": True,
                "description": "Cuota mensual para actividades dirigidas (Zumba, Fitness, Pilates)",
                "is_recurring": True, "recurring_period": "mensual", "recurring_amount": 25.00,
                "created_at": now_iso,
            },
        ]
        await db.products.insert_many(products)
        logger.info("Default products seeded")


# ── RGPD / PROTECCIÓN DE DATOS ──────────────────────────────────────────────

@api_router.get("/gdpr/consents")
async def gdpr_list_consents(request: Request):
    """Lista todos los interesados con su estado de consentimiento RGPD."""
    await get_current_user(request)
    players = await db.players.find({}, {"_id": 0, "id": 1, "name": 1, "surname": 1, "email": 1,
        "consent_gdpr": 1, "consent_communications": 1, "consent_date": 1, "consent_ip": 1,
        "consent_guardian_gdpr": 1, "created_at": 1}).to_list(2000)
    members = await db.members.find({}, {"_id": 0, "id": 1, "name": 1, "surname": 1, "email": 1,
        "consent_gdpr": 1, "consent_communications": 1, "consent_date": 1, "consent_ip": 1,
        "created_at": 1}).to_list(2000)
    return {
        "players": [{"type": "player", **p} for p in players],
        "members": [{"type": "member", **m} for m in members],
    }

@api_router.get("/gdpr/export/{person_type}/{person_id}")
async def gdpr_export_person(person_type: str, person_id: str, request: Request):
    """
    Exporta todos los datos de una persona (RGPD Art. 15 — Derecho de acceso,
    Art. 20 — Portabilidad). Devuelve JSON con todos los registros asociados.
    """
    await get_current_user(request)
    if person_type not in ("player", "member"):
        raise HTTPException(status_code=400, detail="Tipo debe ser 'player' o 'member'")

    col = db.players if person_type == "player" else db.members
    person = await col.find_one({"id": person_id}, {"_id": 0})
    if not person:
        raise HTTPException(status_code=404, detail="Persona no encontrada")

    # Enmascarar datos sensibles en la exportación (IBAN)
    if person.get("bank_iban"):
        person["bank_iban"] = mask_iban(person["bank_iban"])

    # Recopilar todos los datos asociados
    payments = await db.payments.find({"person_id": person_id}, {"_id": 0}).to_list(200)
    sepa_mandates = await db.sepa_mandates.find({"person_id": person_id}, {"_id": 0}).to_list(50)
    for m in sepa_mandates:
        m["debtor_iban"] = mask_iban(m.get("debtor_iban", ""))

    guardians = []
    if person_type == "player":
        family_id = person.get("family_id")
        if family_id:
            guardians = await db.guardians.find({"family_id": family_id}, {"_id": 0}).to_list(10)
            for g in guardians:
                if g.get("bank_iban"):
                    g["bank_iban"] = mask_iban(g["bank_iban"])

    export = {
        "exportacion_rgpd": {
            "fecha": datetime.now(timezone.utc).isoformat(),
            "responsable": "El club solicitante",
            "encargado": "SUDEPORTE (gescomsport@gmail.com)",
            "base_juridica": "Art. 6.1.b RGPD — Ejecución de contrato",
            "derechos": "Puedes ejercer tus derechos ARCO contactando al club.",
        },
        "datos_personales": person,
        "pagos": payments,
        "mandatos_sepa": sepa_mandates,
        "tutores_legales": guardians,
    }

    await db.gdpr_log.insert_one({
        "id": str(uuid.uuid4()),
        "action": "export",
        "person_type": person_type,
        "person_id": person_id,
        "date": datetime.now(timezone.utc).isoformat(),
        "requested_by": "admin",
    })

    return Response(
        content=json.dumps(export, ensure_ascii=False, indent=2),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=datos_rgpd_{person_id}.json"}
    )

@api_router.delete("/gdpr/delete/{person_type}/{person_id}")
async def gdpr_delete_person(person_type: str, person_id: str, request: Request):
    """
    Elimina todos los datos de una persona (RGPD Art. 17 — Derecho al olvido).
    Conserva únicamente el registro mínimo para cumplimiento fiscal (Art. 17.3.b).
    """
    await get_current_user(request)
    if person_type not in ("player", "member"):
        raise HTTPException(status_code=400, detail="Tipo debe ser 'player' o 'member'")

    col = db.players if person_type == "player" else db.members
    person = await col.find_one({"id": person_id}, {"_id": 0})
    if not person:
        raise HTTPException(status_code=404, detail="Persona no encontrada")

    now_iso = datetime.now(timezone.utc).isoformat()

    # Anonimizar en lugar de borrar: conservar mínimo fiscal
    anonymized = {
        "id": person_id,
        "name": "ANONIMIZADO",
        "surname": "ANONIMIZADO",
        "email": "",
        "phone": "",
        "dni": "",
        "address": "",
        "city": "",
        "postal_code": "",
        "bank_iban": "",
        "photo_url": "",
        "status": "deleted_gdpr",
        "gdpr_deletion_date": now_iso,
        "notes": f"Datos eliminados por solicitud RGPD el {now_iso[:10]}",
    }
    await col.update_one({"id": person_id}, {"$set": anonymized})

    # Cancelar mandatos SEPA
    await db.sepa_mandates.update_many(
        {"person_id": person_id},
        {"$set": {"status": "cancelled", "debtor_iban": "", "person_name": "ANONIMIZADO",
                  "gdpr_cancelled": True, "gdpr_cancelled_date": now_iso}}
    )

    # Registrar la eliminación en log de auditoría
    await db.gdpr_log.insert_one({
        "id": str(uuid.uuid4()),
        "action": "deletion",
        "person_type": person_type,
        "person_id": person_id,
        "date": now_iso,
        "requested_by": "admin",
    })

    return {"message": f"Datos de {person_type} {person_id} eliminados conforme al RGPD Art. 17."}

@api_router.get("/gdpr/log")
async def gdpr_get_log(request: Request):
    """Registro de auditoría RGPD (accesos, exportaciones, eliminaciones)."""
    await get_current_user(request)
    logs = await db.gdpr_log.find({}, {"_id": 0}).sort("date", -1).to_list(500)
    return logs

@api_router.get("/gdpr/privacy-policy")
async def get_privacy_policy():
    """Política de privacidad del club (pública)."""
    settings = await db.settings.find_one({"type": "club"}, {"_id": 0}) or {}
    return {"privacy_policy": settings.get("privacy_policy", ""), "club_name": settings.get("club_name", "")}


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
